"""工作簿 → 緊湊文字表示（SheetCompressor 簡化版）。

設計目標：
- 保留儲存格座標資訊（列號 + 欄字母），模型才能產生精準操作
- 只輸出實際使用範圍；過長的表格取頭尾抽樣
- 非預設格式（粗體、底色、數值格式）與合併儲存格以附註列出，而非逐格展開
"""
from __future__ import annotations

import datetime as _dt
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


def _render_value(v) -> str:
    if v is None:
        return ""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, _dt.datetime):
        if v.hour == 0 and v.minute == 0 and v.second == 0:
            return v.date().isoformat()
        return v.isoformat(sep=" ")
    if isinstance(v, _dt.date):
        return v.isoformat()
    if isinstance(v, float):
        if v == int(v) and abs(v) < 1e15:
            return str(int(v))
        return f"{v:g}"
    return str(v).replace("|", "¦").replace("\n", "⏎")


def _used_range(ws: Worksheet) -> tuple[int, int]:
    """回傳 (max_row, max_col)，以「實際有值」為準（openpyxl 的 max_row 會被純樣式格灌水）。"""
    max_r = 0
    max_c = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and cell.value != "":
                if cell.row > max_r:
                    max_r = cell.row
                if cell.column > max_c:
                    max_c = cell.column
    return max_r, max_c


def _rgb_tail(color) -> str | None:
    """取 openpyxl 顏色物件的 RGB 後六碼；取不到（theme 色等）回傳 None。"""
    try:
        rgb = color.rgb
    except AttributeError:
        return None
    if not isinstance(rgb, str) or len(rgb) < 6:
        return None
    return rgb[-6:].upper()


def _row_spans(cols: list[int], row: int) -> list[str]:
    """把同一列上的欄號集合壓成 'A3:D3' 這種區段表示。"""
    spans = []
    cols = sorted(cols)
    start = prev = cols[0]
    for c in cols[1:] + [None]:
        if c is not None and c == prev + 1:
            prev = c
            continue
        a = f"{get_column_letter(start)}{row}"
        b = f"{get_column_letter(prev)}{row}"
        spans.append(a if a == b else f"{a}:{b}")
        if c is not None:
            start = prev = c
    return spans


def _collect_format_notes(ws: Worksheet, max_row: int, max_col: int, cap: int = 12) -> list[str]:
    notes: list[str] = []
    scan_rows = min(max_row, 60)

    bold_by_row: dict[int, list[int]] = {}
    fill_by_row: dict[tuple[int, str], list[int]] = {}
    red_by_row: dict[int, list[int]] = {}
    numfmt_by_col: dict[int, dict[str, int]] = {}

    for row in ws.iter_rows(min_row=1, max_row=scan_rows, max_col=max_col):
        for cell in row:
            if cell.font is not None and cell.font.bold:
                bold_by_row.setdefault(cell.row, []).append(cell.column)
            if cell.font is not None:
                frgb = _rgb_tail(cell.font.color) if cell.font.color else None
                if frgb and frgb not in ("000000",):
                    red_by_row.setdefault(cell.row, []).append(cell.column)
            if cell.fill is not None and getattr(cell.fill, "fill_type", None) == "solid":
                rgb = _rgb_tail(cell.fill.start_color)
                if rgb and rgb not in ("000000", "FFFFFF"):
                    fill_by_row.setdefault((cell.row, rgb), []).append(cell.column)
            if cell.row >= 2 and cell.value is not None:
                fmt = cell.number_format or "General"
                if fmt not in ("General", "@"):
                    numfmt_by_col.setdefault(cell.column, {})
                    numfmt_by_col[cell.column][fmt] = numfmt_by_col[cell.column].get(fmt, 0) + 1

    for r, cols in sorted(bold_by_row.items()):
        notes.append("粗體 " + ",".join(_row_spans(cols, r)))
    for (r, rgb), cols in sorted(fill_by_row.items()):
        notes.append(f"底色#{rgb} " + ",".join(_row_spans(cols, r)))
    for r, cols in sorted(red_by_row.items()):
        # 字色附註需要具體色碼，逐列取第一格的顏色
        c0 = ws.cell(row=r, column=cols[0])
        rgb = _rgb_tail(c0.font.color) or "??????"
        notes.append(f"字色#{rgb} " + ",".join(_row_spans(cols, r)))
    for col, fmts in sorted(numfmt_by_col.items()):
        fmt, cnt = max(fmts.items(), key=lambda kv: kv[1])
        notes.append(f"{get_column_letter(col)}欄數值格式「{fmt}」({cnt}格)")

    if ws.merged_cells.ranges:
        merged = ",".join(str(r) for r in list(ws.merged_cells.ranges)[:8])
        notes.append(f"合併儲存格 {merged}")

    if len(notes) > cap:
        notes = notes[:cap] + [f"…(格式附註共 {len(notes)} 項，僅列前 {cap} 項)"]
    return notes


def _detect_header_row(ws: Worksheet, max_row: int, max_col: int) -> int:
    """推測表頭所在列（真實檔案常有標題列在表頭上方）。回傳 1-based 列號。"""
    counts = []
    for r in range(1, min(6, max_row) + 1):
        n = sum(1 for c in range(1, max_col + 1)
                if ws.cell(row=r, column=c).value not in (None, ""))
        counts.append(n)
    best_r, best_n = 1, counts[0]
    for i, n in enumerate(counts):
        if n >= 2 and n > best_n + 1:      # 明顯比上方更「滿」的列才視為表頭
            best_r, best_n = i + 1, n
            break
    return best_r


def _detect_two_tier_header(ws: Worksheet, max_col: int):
    """偵測「表頭跨兩列」：左半欄名在第 1 列，右半欄名在第 2 列（第 1 列是合併群組標題）。

    為什麼需要：真實 BOM（AVTC 檔）就是這種版型——A~G 的欄名在第 1 列、H~R 在第 2 列。
    原本 _detect_header_row 只能挑「一列」，會回報「標頭在第 2 列（第 1 列是標題/說明）」，
    那是錯的，而且模型會照做：headers = ws[2] → 找不到 Find Number / Qty → KeyError。
    實測兩題真實 BOM 任務因此失敗，責任在編碼器不在模型。

    回傳 (左半最後一欄, 群組標題列表) 或 None。
    """
    if ws.max_row < 3:
        return None
    merged_top = [m for m in ws.merged_cells.ranges if m.min_row == 1 and m.max_col > m.min_col]
    if not merged_top:                       # 沒有跨欄的群組標題就不是這種版型
        return None
    merged_cols = {c for m in merged_top for c in range(m.min_col, m.max_col + 1)}
    # 左半 = 從 A 往右，直到碰到空欄或群組標題欄為止。
    # （不能用 min(merged.min_col) - 1：真實檔的左半與群組之間常隔一個空欄，
    #   AVTC 就是 A~G 有欄名、H 空白、合併群組從 I 開始。）
    split = 0
    for c in range(1, max_col + 1):
        if c in merged_cols or ws.cell(1, c).value in (None, ""):
            break
        split = c
    if split < 1:
        return None
    right_named = sum(1 for c in range(split + 1, max_col + 1)
                      if ws.cell(2, c).value not in (None, ""))
    if right_named < 2:
        return None
    groups = [(get_column_letter(m.min_col), get_column_letter(m.max_col),
               ws.cell(1, m.min_col).value) for m in sorted(merged_top, key=lambda x: x.min_col)]
    return split, groups


def encode_sheet(ws: Worksheet, max_rows: int = 40) -> str:
    max_row, max_col = _used_range(ws)
    if max_row == 0:
        return f"【工作表：{ws.title}】(空白)"

    lines = [f"【工作表：{ws.title}】範圍 A1:{get_column_letter(max_col)}{max_row}"]
    two_tier = _detect_two_tier_header(ws, max_col)
    if two_tier:
        split, groups = two_tier
        gtxt = "、".join(f"{a}~{b}「{t}」" for a, b, t in groups if t)
        left, right = get_column_letter(split), get_column_letter(split + 1)
        last = get_column_letter(max_col)
        # 措辭要逐欄講死。第一版寫成「第 1 列該處是群組標題」，模型把左右兩半對調了，
        # 兩題真實 BOM 任務照樣失敗——歧義的提示跟錯誤的提示一樣糟。
        lines.append(
            f"[版型] 注意：表頭跨兩列，左右兩半的取法不同——"
            f"A~{left} 欄的欄名 = 第 1 列的值；"
            f"{right}~{last} 欄的欄名 = 第 2 列的值。"
            f"（{right}~{last} 在第 1 列是跨欄合併的群組標題：{gtxt}，那不是欄名；"
            f"A~{left} 在第 2 列的值也不是欄名。）資料自第 3 列起。")
    else:
        header_row = _detect_header_row(ws, max_row, max_col)
        if header_row > 1:
            lines.append(f"[版型] 注意：欄位標頭疑似在第 {header_row} 列"
                         f"（第 1~{header_row - 1} 列是標題/說明），資料自第 {header_row + 1} 列起。"
                         "讀取欄位時請以標頭列的實際欄位位置為準。")
    lines.append("  " + " | ".join(get_column_letter(c) for c in range(1, max_col + 1)))

    if max_row <= max_rows:
        show = list(range(1, max_row + 1))
        gap_at = None
    else:
        head = max_rows - 6
        show = list(range(1, head + 1)) + list(range(max_row - 4, max_row + 1))
        gap_at = head

    for r in show:
        vals = [_render_value(ws.cell(row=r, column=c).value) for c in range(1, max_col + 1)]
        lines.append(f"{r}| " + " | ".join(vals))
        if gap_at is not None and r == gap_at:
            lines.append(f"…(中略 {max_row - max_rows + 1} 列)…")

    notes = _collect_format_notes(ws, max_row, max_col)
    if notes:
        lines.append("[格式] " + "；".join(notes))
    return "\n".join(lines)


def encode_workbook(path: str | Path, max_rows_per_sheet: int = 40,
                    max_chars: int = 7000) -> str:
    """編碼整本工作簿；超過 max_chars 時自動減少每張表顯示的列數。

    多工作表的真實檔案容易撐爆模型 context（進而讓生成失控），
    因此這裡設總量預算，寧可少顯示幾列也要留足夠空間給模型作答。
    """
    wb = openpyxl.load_workbook(path, data_only=False)
    try:
        for rows in (max_rows_per_sheet, 24, 14, 8):
            parts = [f"工作表清單: {', '.join(wb.sheetnames)}"]
            for name in wb.sheetnames:
                parts.append(encode_sheet(wb[name], max_rows=rows))
            text = "\n\n".join(parts)
            if len(text) <= max_chars:
                return text
        return text
    finally:
        wb.close()
