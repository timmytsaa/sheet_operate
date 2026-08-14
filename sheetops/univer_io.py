"""xlsx ⇄ Univer 快照（IWorkbookData JSON）轉換。

伺服器權威架構：xlsx 永遠在後端（openpyxl 讀寫），Univer 前端只吃這裡產生的
JSON 視圖；使用者在網格的編輯以事件回傳，由 apply_edit 寫回 xlsx。
轉換範圍（夠用就好）：值、粗體、底色、字色、數值格式；日期輸出 ISO 字串。
"""
from __future__ import annotations

import datetime as _dt
from pathlib import Path

import openpyxl

from .encoder import _rgb_tail, _used_range


def _cell_value(v):
    if v is None:
        return None
    if isinstance(v, _dt.datetime):
        if v.hour == 0 and v.minute == 0 and v.second == 0:
            return v.date().isoformat()
        return v.isoformat(sep=" ")
    if isinstance(v, _dt.date):
        return v.isoformat()
    if isinstance(v, (int, float, str, bool)):
        return v
    return str(v)


def workbook_to_snapshot(path: str | Path, pad_rows: int = 20, pad_cols: int = 6) -> dict:
    wb = openpyxl.load_workbook(path, data_only=False)
    styles: dict[str, dict] = {}
    style_ids: dict[tuple, str] = {}

    def style_id(cell) -> str | None:
        bold = bool(cell.font and cell.font.bold)
        fill = None
        if cell.fill is not None and getattr(cell.fill, "fill_type", None) == "solid":
            rgb = _rgb_tail(cell.fill.start_color)
            if rgb and rgb not in ("000000", "FFFFFF"):
                fill = rgb
        color = None
        if cell.font is not None and cell.font.color is not None:
            rgb = _rgb_tail(cell.font.color)
            if rgb and rgb != "000000":
                color = rgb
        fmt = cell.number_format if cell.number_format not in (None, "General") else None
        key = (bold, fill, color, fmt)
        if key == (False, None, None, None):
            return None
        if key not in style_ids:
            s: dict = {}
            if bold:
                s["bl"] = 1
            if fill:
                s["bg"] = {"rgb": f"#{fill}"}
            if color:
                s["cl"] = {"rgb": f"#{color}"}
            if fmt:
                s["n"] = {"pattern": fmt}
            sid = f"s{len(style_ids) + 1}"
            style_ids[key] = sid
            styles[sid] = s
        return style_ids[key]

    sheets: dict[str, dict] = {}
    order: list[str] = []
    for idx, name in enumerate(wb.sheetnames):
        ws = wb[name]
        max_r, max_c = _used_range(ws)
        cell_data: dict = {}
        for row in ws.iter_rows(min_row=1, max_row=max_r, max_col=max_c):
            for cell in row:
                v = _cell_value(cell.value)
                sid = style_id(cell)
                if v is None and sid is None:
                    continue
                entry: dict = {}
                if v is not None:
                    entry["v"] = v
                if sid:
                    entry["s"] = sid
                cell_data.setdefault(str(cell.row - 1), {})[str(cell.column - 1)] = entry
        sheet_id = f"sheet{idx + 1}"
        order.append(sheet_id)
        sheets[sheet_id] = {
            "id": sheet_id, "name": name,
            "rowCount": max(max_r + pad_rows, 40),
            "columnCount": max(max_c + pad_cols, 12),
            "cellData": cell_data,
        }
    wb.close()
    return {"id": "workbook", "name": Path(path).stem, "locale": "zhCN",
            "styles": styles, "sheetOrder": order, "sheets": sheets}


def apply_edit(path: str | Path, sheet: str, row: int, col: int, value) -> bool:
    """把前端的儲存格編輯（0-based row/col）寫回 xlsx。回傳是否成功。"""
    wb = openpyxl.load_workbook(path)
    if sheet not in wb.sheetnames:
        wb.close()
        return False
    if value == "":
        value = None
    wb[sheet].cell(row=row + 1, column=col + 1, value=value)
    wb.save(path)
    wb.close()
    return True
