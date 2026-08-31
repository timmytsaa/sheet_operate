"""比對兩個工作簿 → 繁中變更摘要。

部署端「預覽後確認」流程的核心：模型改完的結果先經此摘要給使用者看，
確認後才寫入。永不讓使用者在看不見變更的情況下覆寫檔案。
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

from .encoder import _used_range
from .verifier import _norm

MAX_SAMPLES = 12
PREVIEW_ROWS = 8      # 新工作表預覽列數


def _fmt(v) -> str:
    if v is None:
        return "(空)"
    s = str(v)
    return s if len(s) <= 24 else s[:21] + "…"


def diff_workbooks(before_path: str | Path, after_path: str | Path) -> dict:
    bwb = openpyxl.load_workbook(before_path, data_only=False)
    awb = openpyxl.load_workbook(after_path, data_only=False)

    result = {
        "sheets_added": [s for s in awb.sheetnames if s not in bwb.sheetnames],
        "sheets_removed": [s for s in bwb.sheetnames if s not in awb.sheetnames],
        "sheets": {},  # name -> {changed, samples, dims_before, dims_after}
    }

    for name in bwb.sheetnames:
        if name not in awb.sheetnames:
            continue
        bws, aws = bwb[name], awb[name]
        b_r, b_c = _used_range(bws)
        a_r, a_c = _used_range(aws)
        max_r, max_c = max(b_r, a_r), max(b_c, a_c)

        changed = 0
        samples: list[str] = []
        coords: list[list[int]] = []      # 0-based [row, col]，供前端高亮
        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                bv = _norm(bws.cell(row=r, column=c).value)
                av = _norm(aws.cell(row=r, column=c).value)
                if bv is None and av is None:
                    continue
                if bv != av and not (
                    isinstance(bv, (int, float)) and isinstance(av, (int, float))
                    and not isinstance(bv, bool) and not isinstance(av, bool)
                    and abs(float(bv) - float(av)) <= 1e-9
                ):
                    changed += 1
                    if len(coords) < 3000:
                        coords.append([r - 1, c - 1])
                    if len(samples) < MAX_SAMPLES:
                        coord = f"{get_column_letter(c)}{r}"
                        samples.append(f"{coord}: {_fmt(bws.cell(row=r, column=c).value)}"
                                       f" → {_fmt(aws.cell(row=r, column=c).value)}")
        if changed or (b_r, b_c) != (a_r, a_c):
            result["sheets"][name] = {
                "changed": changed, "samples": samples, "coords": coords,
                "dims_before": f"{b_r}列×{b_c}欄", "dims_after": f"{a_r}列×{a_c}欄",
            }

    # 新工作表要給內容，不能只報名字。這個模型的失敗模式是「看起來正常的錯答案」：
    # 實測真實 BOM 時，一題該輸出 6 列卻給了 66 列、一題漏掉整類差異——
    # 兩次的預覽都只有「＋ 新增工作表「X」」一行，使用者根本無從發現。
    result["added_preview"] = {}
    for name in result["sheets_added"]:
        ws = awb[name]
        n_r, n_c = _used_range(ws)
        rows = [[_fmt(ws.cell(row=r, column=c).value) for c in range(1, min(n_c, 8) + 1)]
                for r in range(1, min(n_r, PREVIEW_ROWS) + 1)]
        result["added_preview"][name] = {
            "rows": n_r, "cols": n_c, "truncated_cols": n_c > 8,
            "head": rows, "more": max(0, n_r - PREVIEW_ROWS),
        }

    bwb.close()
    awb.close()
    return result


def render_diff(d: dict) -> str:
    lines: list[str] = []
    for s in d["sheets_added"]:
        p = (d.get("added_preview") or {}).get(s)
        if not p:
            lines.append(f"＋ 新增工作表「{s}」")
            continue
        lines.append(f"＋ 新增工作表「{s}」　{p['rows']} 列 × {p['cols']} 欄")
        for row in p["head"]:
            cells = " | ".join((v if len(v) <= 18 else v[:17] + "…") for v in row)
            lines.append(f"    {cells}" + ("  …" if p["truncated_cols"] else ""))
        if p["more"]:
            lines.append(f"    …（其餘 {p['more']} 列）")
    for s in d["sheets_removed"]:
        lines.append(f"－ 刪除工作表「{s}」")
    for name, info in d["sheets"].items():
        head = f"◆ 工作表「{name}」：{info['changed']} 個儲存格變更"
        if info["dims_before"] != info["dims_after"]:
            head += f"（範圍 {info['dims_before']} → {info['dims_after']}）"
        lines.append(head)
        for s in info["samples"]:
            lines.append(f"    {s}")
        if info["changed"] > len(info["samples"]):
            lines.append(f"    …（其餘 {info['changed'] - len(info['samples'])} 處省略）")
    if not lines:
        lines.append("（沒有偵測到儲存格值的變更——可能只有格式調整）")
    return "\n".join(lines)
