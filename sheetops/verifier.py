"""比對「執行結果」與「目標工作簿」→ 計算 reward。

check 設定（task.json 的 "check" 欄位）：
{
  "target_sheets": ["訂單"],          # 省略時檢查目標檔的所有工作表
  "float_tol": 1e-6,
  "format_checks": [                  # 選填；格式類任務使用
    {"sheet": "訂單", "range": "A1:F1",
     "props": {"bold": true, "fill_rgb": "D9E1F2", "font_rgb": "FF0000",
               "number_format": "#,##0"}}
  ]
}

回傳：
{score, full_match, value_total, value_match, format_total, format_match, mismatches[...]}
score = 匹配格數 / 總檢查數（值與格式合計），full_match = 全數通過。
"""
from __future__ import annotations

import datetime as _dt
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter, range_boundaries

from .encoder import _rgb_tail, _used_range

MAX_MISMATCH_DETAIL = 20


def _norm(v):
    """正規化儲存格值以便比較。None/空字串視為空；午夜的 datetime 視為 date。"""
    if v is None or (isinstance(v, str) and v == ""):
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, _dt.datetime):
        if v.hour == 0 and v.minute == 0 and v.second == 0 and v.microsecond == 0:
            return v.date()
        return v
    if isinstance(v, str):
        return v
    return v


def _num_eq(a, b, tol: float) -> bool:
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def _values_equal(rv, gv, tol: float) -> bool:
    rv, gv = _norm(rv), _norm(gv)
    if rv is None and gv is None:
        return True
    if isinstance(rv, bool) or isinstance(gv, bool):
        return rv is gv if isinstance(rv, bool) and isinstance(gv, bool) else False
    if isinstance(rv, (int, float)) and isinstance(gv, (int, float)):
        return _num_eq(rv, gv, tol)
    return rv == gv


class _FormulaValues:
    """結果檔含公式時的後援：用 `formulas` 套件求值（選用依賴，失敗則放棄）。"""

    def __init__(self, path: str):
        self.path = path
        self._values: dict[tuple[str, str], object] | None = None
        self.error: str | None = None

    def get(self, sheet: str, coord: str):
        if self._values is None:
            self._values = {}
            try:
                import formulas  # type: ignore
                model = formulas.ExcelModel().loads(self.path).finish()
                sol = model.calculate()
                fname = Path(self.path).name.upper()
                for key, val in sol.items():
                    # key 形如 "'[FILE.XLSX]工作表'!A1"
                    if not (key.startswith("'[") and "]" in key and "'!" in key):
                        continue
                    fpart, rest = key[2:].split("]", 1)
                    if fpart.upper() != fname:
                        continue
                    sname, coord_part = rest.split("'!", 1)
                    try:
                        v = val.value[0, 0] if hasattr(val, "value") else val
                    except Exception:
                        v = None
                    self._values[(sname, coord_part)] = v
            except Exception as e:  # 套件未安裝或求值失敗
                self.error = f"{type(e).__name__}: {e}"
        return self._values.get((sheet, coord.upper()))


def verify(result_path: str | Path, goal_path: str | Path, check: dict | None = None) -> dict:
    check = check or {}
    tol = float(check.get("float_tol", 1e-6))
    result_path, goal_path = str(result_path), str(goal_path)

    report = {
        "score": 0.0, "full_match": False,
        "value_total": 0, "value_match": 0,
        "format_total": 0, "format_match": 0,
        "mismatches": [],
    }

    try:
        rwb = openpyxl.load_workbook(result_path, data_only=False)
    except Exception as e:
        report["mismatches"].append(f"結果檔無法開啟: {type(e).__name__}: {e}")
        return report
    gwb = openpyxl.load_workbook(goal_path, data_only=False)

    formula_vals = _FormulaValues(result_path)

    def add_mismatch(msg: str):
        if len(report["mismatches"]) < MAX_MISMATCH_DETAIL:
            report["mismatches"].append(msg)

    # ---- 值比對 ----
    sheets = check.get("target_sheets") or gwb.sheetnames
    for sname in sheets:
        gws = gwb[sname]
        g_r, g_c = _used_range(gws)
        if sname not in rwb.sheetnames:
            cells = sum(1 for row in gws.iter_rows(max_row=g_r, max_col=g_c)
                        for c in row if c.value is not None)
            report["value_total"] += max(cells, 1)
            add_mismatch(f"缺少工作表「{sname}」")
            continue
        rws = rwb[sname]
        r_r, r_c = _used_range(rws)
        max_r, max_c = max(g_r, r_r), max(g_c, r_c)

        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                gv = gws.cell(row=r, column=c).value
                rv = rws.cell(row=r, column=c).value
                if _norm(gv) is None and _norm(rv) is None:
                    continue
                report["value_total"] += 1
                coord = f"{get_column_letter(c)}{r}"

                if isinstance(rv, str) and rv.startswith("="):
                    ev = formula_vals.get(sname, coord)
                    if ev is not None and _values_equal(ev, gv, tol):
                        report["value_match"] += 1
                        continue
                    if formula_vals.error:
                        add_mismatch(f"{sname}!{coord} 為公式 {rv}，無法求值"
                                     f"（{formula_vals.error}）")
                    else:
                        add_mismatch(f"{sname}!{coord} 公式求值 {ev!r} ≠ 目標 {gv!r}")
                    continue

                if _values_equal(rv, gv, tol):
                    report["value_match"] += 1
                else:
                    add_mismatch(f"{sname}!{coord} 值 {rv!r} ≠ 目標 {gv!r}")

    # ---- 格式比對 ----
    for fc in check.get("format_checks", []):
        sname = fc["sheet"]
        props: dict = fc["props"]
        if sname not in rwb.sheetnames:
            report["format_total"] += 1
            add_mismatch(f"格式檢查失敗：缺少工作表「{sname}」")
            continue
        rws = rwb[sname]
        c1, r1, c2, r2 = range_boundaries(fc["range"])
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                cell = rws.cell(row=r, column=c)
                coord = f"{get_column_letter(c)}{r}"
                for prop, want in props.items():
                    report["format_total"] += 1
                    ok = False
                    if prop == "bold":
                        ok = bool(cell.font and cell.font.bold) == bool(want)
                    elif prop == "fill_rgb":
                        got = _rgb_tail(cell.fill.start_color) if (
                            cell.fill and getattr(cell.fill, "fill_type", None) == "solid") else None
                        ok = (got == str(want).upper())
                    elif prop == "font_rgb":
                        got = _rgb_tail(cell.font.color) if (cell.font and cell.font.color) else None
                        ok = (got == str(want).upper())
                    elif prop == "number_format":
                        ok = (cell.number_format == want)
                    elif prop == "grid_border":
                        has = cell.border is not None and all(
                            getattr(cell.border, side) is not None and
                            getattr(cell.border, side).style
                            for side in ("left", "right", "top", "bottom"))
                        ok = has == bool(want)
                    elif prop == "align_h":
                        got = cell.alignment.horizontal if cell.alignment else None
                        ok = (got == want)
                    else:
                        add_mismatch(f"未知格式屬性 {prop}")
                    if ok:
                        report["format_match"] += 1
                    else:
                        add_mismatch(f"{sname}!{coord} 格式 {prop} 不符（要求 {want!r}）")

    # ---- 新工作表比對（模型自己命名時用：只看內容不看名字） ----
    ns = check.get("new_sheet")
    if ns:
        known = set(ns.get("known_sheets", []))
        added = [s for s in rwb.sheetnames if s not in known]
        report["format_total"] += 1
        if len(added) != 1:
            add_mismatch(f"應新增剛好 1 張工作表，實際新增 {len(added)} 張：{added}")
        else:
            report["format_match"] += 1
            ws = rwb[added[0]]
            headers = ns.get("headers") or []
            rows = ns.get("rows") or []
            # 標題列可有可無：對得上就從第 2 列比資料，對不上則視為缺標題（扣分但仍比資料）
            first = [ws.cell(row=1, column=c + 1).value for c in range(len(headers))]
            has_header = headers and all(
                _values_equal(first[c], headers[c], tol) for c in range(len(headers)))
            if headers:
                report["value_total"] += 1
                if has_header:
                    report["value_match"] += 1
                else:
                    add_mismatch(f"新工作表「{added[0]}」缺少標題列（應為 {headers}）")
            offset = 2 if has_header else 1
            for i, exp_row in enumerate(rows):
                for j, exp_v in enumerate(exp_row):
                    report["value_total"] += 1
                    got = ws.cell(row=offset + i, column=j + 1).value
                    if _values_equal(got, exp_v, tol):
                        report["value_match"] += 1
                    else:
                        coord = f"{get_column_letter(j + 1)}{offset + i}"
                        add_mismatch(f"新工作表「{added[0]}」{coord} 為 {got!r}，應為 {exp_v!r}")
            # 多出來的資料列也算錯
            extra = _used_range(ws)[0] - (offset + len(rows) - 1)
            report["value_total"] += 1
            if extra <= 0:
                report["value_match"] += 1
            else:
                add_mismatch(f"新工作表「{added[0]}」多出 {extra} 列（應只有 {len(rows)} 筆資料）")

    # ---- 公式存在性檢查（值對但填死值者不算通過） ----
    for fc in check.get("formula_cells", []):
        sname = fc["sheet"]
        want_fn = (fc.get("contains") or "").upper()
        if sname not in rwb.sheetnames:
            report["format_total"] += 1
            add_mismatch(f"公式檢查失敗：缺少工作表「{sname}」")
            continue
        rws = rwb[sname]
        c1, r1, c2, r2 = range_boundaries(fc["range"])
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                report["format_total"] += 1
                v = rws.cell(row=r, column=c).value
                coord = f"{get_column_letter(c)}{r}"
                if not (isinstance(v, str) and v.startswith("=")):
                    add_mismatch(f"{sname}!{coord} 必須是公式（目前為 {v!r}）")
                elif want_fn and want_fn not in v.upper():
                    add_mismatch(f"{sname}!{coord} 公式需使用 {want_fn}（目前為 {v!r}）")
                else:
                    report["format_match"] += 1

    # ---- 合併儲存格比對 ----
    for mc in check.get("merged", []):
        sname = mc["sheet"]
        report["format_total"] += 1
        if sname not in rwb.sheetnames:
            add_mismatch(f"合併檢查失敗：缺少工作表「{sname}」")
            continue
        merged_set = {str(r) for r in rwb[sname].merged_cells.ranges}
        if mc["range"].upper() in merged_set:
            report["format_match"] += 1
        else:
            add_mismatch(f"{sname} 缺少合併儲存格 {mc['range']}（現有：{sorted(merged_set)}）")

    rwb.close()
    gwb.close()

    total = report["value_total"] + report["format_total"]
    match = report["value_match"] + report["format_match"]
    report["score"] = (match / total) if total else 0.0
    report["full_match"] = total > 0 and match == total
    return report
