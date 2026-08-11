"""任務產生器共用基底：TaskSpec、工作簿寫入工具、schema 定義與資料表產生器。

schema 多樣化：同樣的任務家族輪換不同「表格外皮」（訂單/報銷/庫存/工時/銷售），
欄位名稱、工作表名稱、資料分佈都不同，但欄位「角色」一致：
  id / person / group / item / cat / qty / price / amount / date
家族程式碼一律透過 meta["col"][role] 取得實際欄名，指令模板用 token 帶入。
"""
from __future__ import annotations

import datetime as _dt
import json
from dataclasses import dataclass, field
from pathlib import Path
from random import Random

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .. import zh_data


@dataclass
class TaskSpec:
    id: str
    family: str
    instruction: str
    check: dict
    ref_solution: str
    instruction_seed: str = ""      # 模板原句（paraphrase 改寫後保留原句於此）
    context: str = ""               # 選配【補充說明】：使用者自訂規則，模型必須遵循才能解對
    meta: dict = field(default_factory=dict)

    def to_json(self) -> str:
        d = dict(self.__dict__)
        if not d["instruction_seed"]:
            d["instruction_seed"] = d["instruction"]
        return json.dumps(d, ensure_ascii=False, indent=2)


# ----------------------------------------------------------------------
# Schema 定義（表格外皮）
# ----------------------------------------------------------------------

SCHEMAS: dict[str, dict] = {
    "orders": {
        "sheet": "訂單", "lookup_sheet": "價目表",
        "id": "訂單編號", "id_prefix": "TW", "id_start": 1001,
        "person": "客戶", "group": "城市", "item": "產品", "cat": "類別",
        "qty": "數量", "price": "單價", "amount": "金額", "date": "訂單日期",
        "groups": zh_data.CITIES, "items": zh_data.PRODUCTS,
        "group_default": "未知",
    },
    "expense": {
        "sheet": "報銷明細", "lookup_sheet": "費用標準",
        "id": "單號", "id_prefix": "EX", "id_start": 2001,
        "person": "申請人", "group": "部門", "item": "費用項目", "cat": "費用類別",
        "qty": "數量", "price": "單價", "amount": "金額", "date": "申請日期",
        "groups": zh_data.DEPARTMENTS, "items": zh_data.EXPENSE_ITEMS,
        "group_default": "未分類",
    },
    "inventory": {
        "sheet": "庫存清單", "lookup_sheet": "單價表",
        "id": "料號", "id_prefix": "MAT", "id_start": 3001,
        "person": "負責人", "group": "倉庫", "item": "品名", "cat": "分類",
        "qty": "數量", "price": "單價", "amount": "庫存金額", "date": "入庫日期",
        "groups": zh_data.WAREHOUSES, "items": zh_data.INVENTORY_ITEMS,
        "group_default": "待分派",
    },
    "timesheet": {
        "sheet": "工時紀錄", "lookup_sheet": "費率表",
        "id": "編號", "id_prefix": "TS", "id_start": 5001,
        "person": "員工", "group": "專案", "item": "工作項目", "cat": "類型",
        "qty": "時數", "price": "費率", "amount": "費用", "date": "日期",
        "groups": zh_data.PROJECTS, "items": zh_data.WORK_ITEMS,
        "group_default": "未指定",
    },
    "sales": {
        "sheet": "銷售紀錄", "lookup_sheet": "定價表",
        "id": "單號", "id_prefix": "SO", "id_start": 7001,
        "person": "業務員", "group": "門市", "item": "商品", "cat": "類別",
        "qty": "數量", "price": "單價", "amount": "銷售額", "date": "銷售日期",
        "groups": zh_data.STORES, "items": zh_data.PRODUCTS,
        "group_default": "總店",
    },
    # 僅評測（OOD）：不進 TRAIN_SCHEMAS，模型訓練時從未見過這套欄位
    "hr": {
        "sheet": "出勤紀錄", "lookup_sheet": "津貼標準",
        "id": "編號", "id_prefix": "HR", "id_start": 8001,
        "person": "員工", "group": "廠區", "item": "勤務項目", "cat": "類型",
        "qty": "時數", "price": "單價", "amount": "小計", "date": "日期",
        "groups": zh_data.PLANTS, "items": zh_data.HR_ITEMS,
        "group_default": "總部",
    },
}

ROLES = ["id", "person", "group", "item", "cat", "qty", "price", "amount", "date"]

# 訓練時輪換的 schema（hr 排除——保留為 OOD 評測專用）
TRAIN_SCHEMAS = sorted(k for k in SCHEMAS if k != "hr")

# gen_tasks --schema 用的全域覆寫；None = 隨機輪換 TRAIN_SCHEMAS
FORCE_SCHEMA: str | None = None


def new_wb() -> Workbook:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    return wb


def write_table(ws: Worksheet, headers: list, rows: list[list],
                style_header: bool = True, date_fmt: str = "yyyy-mm-dd") -> None:
    ws.append(headers)
    if style_header:
        for cell in ws[1]:
            cell.font = Font(bold=True)
    for r in rows:
        ws.append(r)
        for cell in ws[ws.max_row]:
            if isinstance(cell.value, (_dt.date, _dt.datetime)):
                cell.number_format = date_fmt
    for ci in range(1, len(headers) + 1):
        width = 8
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=ci).value
            if v is None:
                continue
            w = sum(2 if ord(ch) > 0x2E80 else 1 for ch in str(v)) + 2
            width = max(width, w)
        ws.column_dimensions[get_column_letter(ci)].width = min(width, 30)


def build_wb(sheets: dict[str, tuple[list, list[list]]], style_header: bool = True) -> Workbook:
    wb = new_wb()
    for name, (headers, rows) in sheets.items():
        write_table(wb.create_sheet(name), headers, rows, style_header=style_header)
    return wb


def write_task(out_root: str | Path, spec: TaskSpec,
               start_wb: Workbook, goal_wb: Workbook) -> Path:
    task_dir = Path(out_root) / spec.id
    task_dir.mkdir(parents=True, exist_ok=True)
    start_wb.save(task_dir / "start.xlsx")
    goal_wb.save(task_dir / "goal.xlsx")
    (task_dir / "task.json").write_text(spec.to_json(), encoding="utf-8")
    return task_dir


def fill(template: str, **kw) -> str:
    """把模板中的 __KEY__ 佔位符替換為值。"""
    for k, v in kw.items():
        template = template.replace(f"__{k.upper()}__", str(v))
    return template


def schema_tokens(s: dict) -> dict:
    """schema → fill() 用的 token 組（sheet/lookup/person/group/... 皆為實際名稱）。"""
    return {
        "sheet": s["sheet"], "lookup": s["lookup_sheet"],
        "id_h": s["id"], "person": s["person"], "group": s["group"],
        "item": s["item"], "cat": s["cat"], "qty": s["qty"],
        "price": s["price"], "amt": s["amount"], "date_h": s["date"],
        "default": s["group_default"],
    }


# ----------------------------------------------------------------------
# 資料表產生器（schema 泛化版）
# ----------------------------------------------------------------------

def make_table(rng: Random, schema: str | dict | None = None, n: int | None = None, *,
               with_price: bool = True, with_amount: bool = True,
               date_col: bool = False, split_name: bool = False,
               empty_price_amount: bool = False, consistent_price: bool = False,
               blank_group: int = 0, padded_names: int = 0, dup_rows: int = 0,
               unique_amounts: bool = False, group_pool: int | None = None):
    """產生一張資料表 (headers, rows, meta)。

    meta:
      s      —— schema dict（含名稱）
      col    —— role → 實際欄名
      idx    —— 欄名 → 欄索引（0-based）
      price_map / groups
    """
    if schema is None:
        schema = FORCE_SCHEMA or rng.choice(TRAIN_SCHEMAS)
    s = SCHEMAS[schema] if isinstance(schema, str) else schema

    if n is None:
        n = rng.randint(12, 18)
    k = min(len(s["groups"]), group_pool or rng.randint(3, 5))
    groups = rng.sample(s["groups"], k)

    price_map: dict[str, int] = {}
    rows: list[list] = []
    seen_amounts: set[int] = set()

    def sample_item():
        name, cat, lo, hi = rng.choice(s["items"])
        price = rng.randrange(lo // 10, hi // 10 + 1) * 10
        if consistent_price:
            price = price_map.setdefault(name, price)
        return name, cat, price

    for i in range(n):
        pname, cat, price = sample_item()
        qty = rng.randint(1, 9)
        if unique_amounts:
            tries = 0
            while qty * price in seen_amounts and tries < 50:
                pname, cat, price = sample_item()
                qty = rng.randint(1, 9)
                tries += 1
            while qty * price in seen_amounts:   # 大表後援：微調單價直到唯一（不動 rng 流）
                price += 10
            seen_amounts.add(qty * price)

        row: list = [f"{s['id_prefix']}{s['id_start'] + i}"]
        if split_name:
            sur, giv = zh_data.split_person_name(rng)
            row += [sur, giv]
        else:
            row.append(zh_data.person_name(rng))
        row.append(rng.choice(groups))
        row += [pname, cat, qty]
        if with_price:
            row.append(None if empty_price_amount else price)
        if with_amount:
            row.append(None if empty_price_amount else qty * price)
        if date_col:
            row.append(zh_data.order_date(rng))
        rows.append(row)
        if not consistent_price:
            price_map.setdefault(pname, price)

    headers: list[str] = [s["id"]]
    headers += ["姓", "名"] if split_name else [s["person"]]
    headers += [s["group"], s["item"], s["cat"], s["qty"]]
    if with_price:
        headers.append(s["price"])
    if with_amount:
        headers.append(s["amount"])
    if date_col:
        headers.append(s["date"])

    idx = {h: i for i, h in enumerate(headers)}
    col = {r: s[r] for r in ROLES}

    # 確保至少兩個類別（彙總類任務需要）
    cat_i = idx[s["cat"]]
    if len({r[cat_i] for r in rows}) < 2:
        cur = rows[-1][cat_i]
        alt = next(t for t in s["items"] if t[1] != cur)
        pname, cat, lo, _hi = alt
        price = (lo // 10) * 10
        if consistent_price:
            price = price_map.setdefault(pname, price)
        rows[-1][idx[s["item"]]] = pname
        rows[-1][cat_i] = cat
        if with_price and not empty_price_amount:
            rows[-1][idx[s["price"]]] = price
            if with_amount:
                rows[-1][idx[s["amount"]]] = rows[-1][idx[s["qty"]]] * price
        price_map.setdefault(pname, price)

    # 資料瑕疵注入（清理類任務用）
    if blank_group:
        for r_i in rng.sample(range(n), min(blank_group, n)):
            rows[r_i][idx[s["group"]]] = None
    if padded_names and not split_name:
        ni = idx[s["person"]]
        for r_i in rng.sample(range(n), min(padded_names, n)):
            rows[r_i][ni] = " " + rows[r_i][ni] + " "
    if dup_rows:
        for _ in range(dup_rows):
            src = rng.randrange(len(rows))
            rows.insert(rng.randrange(src + 1, len(rows) + 1), list(rows[src]))

    meta = {"s": s, "schema": s["sheet"], "col": col, "idx": idx,
            "price_map": price_map, "groups": groups, "n": len(rows)}
    return headers, rows, meta
