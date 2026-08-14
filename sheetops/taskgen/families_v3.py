"""v3 任務家族：非標準版型的座標對地（源自第一個真實失敗案例）。

真實檔案的表格常有：標題列壓在表頭上方、表頭不在第 1 列、表格不從 A 欄開始、
標題與表頭之間夾空白列。訓練資料 v1/v2 的表頭 100% 在第 1 列 A 欄起——這個盲區
讓部署後的第一個真實任務連續兩次交白卷（含欄位索引整體位移一格的失誤）。

編碼器已會輸出「[版型] 表頭疑似在第 N 列」註記；本家族讓模型學會使用它。
"""
from __future__ import annotations

import datetime as _dt
from random import Random

from openpyxl.styles import Font

from .base import TaskSpec, build_wb, fill, make_table, new_wb, schema_tokens
from .families import _ai

_TITLES = ["{sheet}明細表", "{sheet}（月結）", "廠內{sheet}統計", "{sheet} 彙整清單"]

_DEFECTS = ["刮傷", "破裂", "髒污", "色差", "鬆動", "缺件", "壓痕"]


def _write_offset(ws, headers, rows, header_row=1, start_col=1, title=None):
    """把表格寫在非標準位置：標題（選）、表頭列、資料列，支援欄位偏移。"""
    if title:
        c = ws.cell(row=1, column=start_col, value=title)
        c.font = Font(bold=True)
    for j, h in enumerate(headers):
        ws.cell(row=header_row, column=start_col + j, value=h).font = Font(bold=True)
    for i, r in enumerate(rows):
        for j, v in enumerate(r):
            cell = ws.cell(row=header_row + 1 + i, column=start_col + j, value=v)
            if isinstance(v, (_dt.date, _dt.datetime)):
                cell.number_format = "yyyy-mm-dd"


def _offset_wb(sheet_name, headers, rows, header_row, start_col, title):
    wb = new_wb()
    _write_offset(wb.create_sheet(sheet_name), headers, rows,
                  header_row=header_row, start_col=start_col, title=title)
    return wb


def gen_offset_layout(rng: Random, task_id: str):
    variant = rng.choice(["title_filter", "gap_sort", "colshift_compute",
                          "both_groupby", "cross_filter"])

    if variant == "cross_filter":
        return _gen_cross_filter(rng, task_id)

    headers, rows, meta = make_table(rng, unique_amounts=True,
                                     with_amount=(variant != "colshift_compute"))
    t = schema_tokens(meta["s"])
    title = rng.choice(_TITLES).format(sheet=t["sheet"])

    if variant == "title_filter":
        header_row, start_col = 2, 1
        ai = _ai(meta)
        amts = sorted(r[ai] for r in rows)
        thr = amts[len(amts) // 2]
        goal_rows = [r for r in rows if r[ai] >= thr]
        instruction = fill(rng.choice([
            "請只保留「__SHEET__」中__AMT__大於或等於 __T__ 的資料列，其餘刪除；表格的標題與欄位列維持原樣。",
            "把「__SHEET__」裡__AMT__低於 __T__ 的資料列刪掉（注意別動到最上面的標題）。",
        ]), t=thr, **t)
        ref = fill('''import openpyxl
wb = openpyxl.load_workbook(INPUT_PATH)
ws = wb["__SHEET__"]
HEADER_ROW = 2
headers = [c.value for c in ws[HEADER_ROW]]
rows = [list(r) for r in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True)]
ai = headers.index("__AMT__")
new_rows = [r for r in rows if r[ai] is not None and r[ai] >= __T__]
if ws.max_row > HEADER_ROW:
    ws.delete_rows(HEADER_ROW + 1, ws.max_row - HEADER_ROW)
for r in new_rows:
    ws.append(r)
wb.save(OUTPUT_PATH)
''', t=thr, **t)

    elif variant == "gap_sort":
        header_row, start_col = 3, 1        # 標題列 + 空白列 + 表頭
        goal_rows = sorted(rows, key=lambda r: r[_ai(meta)], reverse=True)
        instruction = fill(rng.choice([
            "請將「__SHEET__」的資料列依「__AMT__」由大到小重新排序（標題與欄位列不動）。",
            "幫「__SHEET__」按__AMT__從高到低排一下，上面的標題那些都保持原位。",
        ]), **t)
        ref = fill('''import openpyxl
wb = openpyxl.load_workbook(INPUT_PATH)
ws = wb["__SHEET__"]
HEADER_ROW = 3
headers = [c.value for c in ws[HEADER_ROW]]
rows = [list(r) for r in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True)]
ai = headers.index("__AMT__")
new_rows = sorted(rows, key=lambda r: r[ai], reverse=True)
if ws.max_row > HEADER_ROW:
    ws.delete_rows(HEADER_ROW + 1, ws.max_row - HEADER_ROW)
for r in new_rows:
    ws.append(r)
wb.save(OUTPUT_PATH)
''', **t)

    elif variant == "colshift_compute":
        header_row, start_col = 1, 3        # 表格從 C 欄開始
        title = None
        qi, pi = meta["idx"][meta["col"]["qty"]], meta["idx"][meta["col"]["price"]]
        goal_headers = headers + [t["amt"]]
        goal_rows = [r + [r[qi] * r[pi]] for r in rows]
        instruction = fill(rng.choice([
            "請在「__SHEET__」表格的最右側新增「__AMT__」欄位＝__QTY__ × __PRICE__（數值）。"
            "注意表格不是從 A 欄開始，請接在實際的最後一欄右邊。",
            "幫「__SHEET__」加一欄「__AMT__」（__QTY__乘__PRICE__），放在表格現有欄位的正右方。",
        ]), **t)
        ref = fill('''import openpyxl
wb = openpyxl.load_workbook(INPUT_PATH)
ws = wb["__SHEET__"]
row1 = [c.value for c in ws[1]]
start = next(i for i, v in enumerate(row1) if v not in (None, ""))  # 0-based 起始欄
headers = [v for v in row1[start:] if v not in (None, "")]
qi = headers.index("__QTY__")
pi = headers.index("__PRICE__")
new_col = start + len(headers) + 1          # 1-based
ws.cell(row=1, column=new_col, value="__AMT__")
r = 2
while ws.cell(row=r, column=start + 1).value is not None:
    q = ws.cell(row=r, column=start + qi + 1).value
    p = ws.cell(row=r, column=start + pi + 1).value
    ws.cell(row=r, column=new_col, value=q * p)
    r += 1
wb.save(OUTPUT_PATH)
''', **t)
        start_wb = _offset_wb(t["sheet"], headers, rows, header_row, start_col, title)
        goal_wb = _offset_wb(t["sheet"], goal_headers, goal_rows, header_row, start_col, title)
        spec = TaskSpec(task_id, "offset_layout", instruction,
                        {"target_sheets": [t["sheet"]]}, ref,
                        meta={"variant": variant, "schema": meta["s"]["sheet"]})
        return spec, start_wb, goal_wb

    else:  # both_groupby：標題列＋表格從 B 欄開始 → 彙總到新表（標準版型）
        header_row, start_col = 2, 2
        ci, ai = meta["idx"][meta["col"]["cat"]], _ai(meta)
        total_h = "總" + t["amt"]
        stats: dict[str, list[int]] = {}
        for r in rows:
            s = stats.setdefault(r[ci], [0, 0])
            s[0] += r[ai]
            s[1] += 1
        sum_rows = [[cat, stats[cat][0], stats[cat][1]] for cat in sorted(stats)]
        instruction = fill(rng.choice([
            "請新增「彙總」工作表（欄位：__CAT__、__TOTAL_H__、筆數，依__CAT__排序），"
            "統計「__SHEET__」中每個__CAT__的__TOTAL_H__與筆數。原表不動。",
            "幫我從「__SHEET__」做分類統計到新工作表「彙總」：各__CAT__一列，"
            "欄位是 __CAT__、__TOTAL_H__、筆數。",
        ]), total_h=total_h, **t)
        ref = fill('''import openpyxl
wb = openpyxl.load_workbook(INPUT_PATH)
ws = wb["__SHEET__"]
HEADER_ROW = 2
row_h = [c.value for c in ws[HEADER_ROW]]
start = next(i for i, v in enumerate(row_h) if v not in (None, ""))
headers = [v for v in row_h[start:] if v not in (None, "")]
ci = headers.index("__CAT__")
ai = headers.index("__AMT__")
stats = {}
for r in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
    vals = list(r)[start:]
    if len(vals) <= max(ci, ai) or vals[ci] is None:
        continue
    s = stats.setdefault(vals[ci], [0, 0])
    s[0] += vals[ai]
    s[1] += 1
out = wb.create_sheet("彙總")
out.append(["__CAT__", "__TOTAL_H__", "筆數"])
for cat in sorted(stats):
    out.append([cat, stats[cat][0], stats[cat][1]])
wb.save(OUTPUT_PATH)
''', total_h=total_h, **t)
        start_wb = _offset_wb(t["sheet"], headers, rows, header_row, start_col, title)
        goal_wb = _offset_wb(t["sheet"], headers, rows, header_row, start_col, title)
        gws = goal_wb.create_sheet("彙總")
        gws.append([t["cat"], total_h, "筆數"])
        for row in sum_rows:
            gws.append(row)
        spec = TaskSpec(task_id, "offset_layout", instruction,
                        {"target_sheets": [t["sheet"], "彙總"]}, ref,
                        meta={"variant": variant, "schema": meta["s"]["sheet"]})
        return spec, start_wb, goal_wb

    start_wb = _offset_wb(t["sheet"], headers, rows, header_row, start_col, title)
    goal_wb = _offset_wb(t["sheet"], headers, goal_rows, header_row, start_col, title)
    spec = TaskSpec(task_id, "offset_layout", instruction,
                    {"target_sheets": [t["sheet"]]}, ref,
                    meta={"variant": variant, "schema": meta["s"]["sheet"]})
    return spec, start_wb, goal_wb


def _gen_cross_filter(rng: Random, task_id: str):
    """OOB 真實案例的複製品：檢驗表（表頭第 2 列）× 主表歸屬過濾 → 問題清單。"""
    n = rng.randint(14, 20)
    groups = ["A組", "B組"]
    ids = [f"SN{rng.randint(10, 99)}{i:03d}" for i in range(n)]
    master_rows = [[i + 1, ids[i], rng.choice(["X100", "X200", "X300"]),
                    rng.choice(groups)] for i in range(n)]
    inspect_rows = []
    for i in range(n):
        status = "ok" if rng.random() < 0.55 else rng.choice(_DEFECTS)
        inspect_rows.append([ids[i], status])
    target = "A組"
    picked = [[r[0], r[1]] for r, m in zip(inspect_rows, master_rows)
              if r[1] != "ok" and m[3] == target]
    if not picked:      # 保底：至少一筆
        for r, m in zip(inspect_rows, master_rows):
            if m[3] == target:
                r[1] = rng.choice(_DEFECTS)
                picked = [[r[0], r[1]]]
                break

    instruction = (f"建立新工作表「{target}問題清單」：從「檢驗紀錄」挑出狀態不是 ok 的資料列，"
                   f"且該編號在「產品清單」的組別欄必須是「{target}」，新表欄位依序為 編號、狀態。")
    ref = f'''import openpyxl
wb = openpyxl.load_workbook(INPUT_PATH)
master = wb["產品清單"]
headers = [c.value for c in master[1]]
id_i = headers.index("編號") + 1
grp_i = headers.index("組別") + 1
group_of = {{}}
for r in range(2, master.max_row + 1):
    sn = master.cell(row=r, column=id_i).value
    if sn is not None:
        group_of[sn] = master.cell(row=r, column=grp_i).value
insp = wb["檢驗紀錄"]
HEADER_ROW = 2
out = wb.create_sheet("{target}問題清單")
out.append(["編號", "狀態"])
for r in range(HEADER_ROW + 1, insp.max_row + 1):
    sn = insp.cell(row=r, column=1).value
    status = insp.cell(row=r, column=2).value
    if sn is None or status is None:
        continue
    if str(status).strip().lower() == "ok":
        continue
    if group_of.get(sn) == "{target}":
        out.append([sn, status])
wb.save(OUTPUT_PATH)
'''
    start_wb = new_wb()
    mws = start_wb.create_sheet("產品清單")
    _write_offset(mws, ["No", "編號", "型號", "組別"], master_rows, header_row=1, start_col=1)
    iws = start_wb.create_sheet("檢驗紀錄")
    _write_offset(iws, ["編號", "狀態"], inspect_rows, header_row=2, start_col=1,
                  title="出廠檢驗紀錄")

    goal_wb = new_wb()
    mws2 = goal_wb.create_sheet("產品清單")
    _write_offset(mws2, ["No", "編號", "型號", "組別"], master_rows, header_row=1, start_col=1)
    iws2 = goal_wb.create_sheet("檢驗紀錄")
    _write_offset(iws2, ["編號", "狀態"], inspect_rows, header_row=2, start_col=1,
                  title="出廠檢驗紀錄")
    pws = goal_wb.create_sheet(f"{target}問題清單")
    _write_offset(pws, ["編號", "狀態"], picked, header_row=1, start_col=1)

    spec = TaskSpec(task_id, "offset_layout", instruction,
                    {"target_sheets": ["產品清單", "檢驗紀錄", f"{target}問題清單"]}, ref,
                    meta={"variant": "cross_filter"})
    return spec, start_wb, goal_wb


V3_FAMILIES = {
    "offset_layout": gen_offset_layout,
}
