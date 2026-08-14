"""v4 任務家族：寫入 Excel 公式（配合 Univer 前端的公式引擎，讓表格是「活」的）。

與其他家族的差異：
- 目標檔存「計算後的值」，模型必須輸出**公式**；驗證器用 formulas 套件求值後比對
- 另有 formula_cells 檢查強制該格真的是公式（填死值不算通過）
- 函數限制在 formulas 套件與 Univer 都支援的常用集：
  SUM / AVERAGE / COUNT / COUNTA / MAX / IF / ROUND / INT / VLOOKUP / SUMIF
"""
from __future__ import annotations

from random import Random

from openpyxl.utils import get_column_letter

from .base import TaskSpec, build_wb, fill, make_table, new_wb, schema_tokens, write_table
from .families import _HDR, _ai

_TOL = {"float_tol": 1e-4}


def gen_formula_write(rng: Random, task_id: str):
    variant = rng.choice(["sum_total", "arith_column", "vlookup_price",
                          "if_grade", "round_tax", "stats_block"])

    if variant == "arith_column":
        headers, rows, meta = make_table(rng, with_amount=False)
        t = schema_tokens(meta["s"])
        qi = meta["idx"][meta["col"]["qty"]]
        pi = meta["idx"][meta["col"]["price"]]
        qL, pL = get_column_letter(qi + 1), get_column_letter(pi + 1)
        newL = get_column_letter(len(headers) + 1)
        last_row = len(rows) + 1
        goal = build_wb({t["sheet"]: (headers + [t["amt"]],
                                      [r + [r[qi] * r[pi]] for r in rows])})
        instruction = fill(rng.choice([
            "請在「__SHEET__」最右側新增「__AMT__」欄，**使用公式** __QTY__ × __PRICE__（例如 =__QL__2*__PL__2），"
            "不要填入計算後的固定值。",
            "幫「__SHEET__」加一欄「__AMT__」，內容要用公式算 __QTY__ 乘 __PRICE__（每列參照該列的儲存格）。",
        ]), ql=qL, pl=pL, **t)
        ref = fill(_HDR + '''from openpyxl.utils import get_column_letter
qL = get_column_letter(headers.index("__QTY__") + 1)
pL = get_column_letter(headers.index("__PRICE__") + 1)
col = len(headers) + 1
ws.cell(row=1, column=col, value="__AMT__")
for i in range(len(rows)):
    r = i + 2
    ws.cell(row=r, column=col, value=f"={qL}{r}*{pL}{r}")
wb.save(OUTPUT_PATH)
''', **t)
        check = {"target_sheets": [t["sheet"]], **_TOL,
                 "formula_cells": [{"sheet": t["sheet"], "range": f"{newL}2:{newL}{last_row}"}]}

    elif variant == "sum_total":
        headers, rows, meta = make_table(rng)
        t = schema_tokens(meta["s"])
        qi, ai = meta["idx"][meta["col"]["qty"]], _ai(meta)
        qL, aL = get_column_letter(qi + 1), get_column_letter(ai + 1)
        last = len(rows) + 1
        total = [None] * len(headers)
        total[0] = "總計"
        total[qi] = sum(r[qi] for r in rows)
        total[ai] = sum(r[ai] for r in rows)
        goal = build_wb({t["sheet"]: (headers, rows + [total])})
        instruction = fill(rng.choice([
            "請在「__SHEET__」資料最底部新增總計列：第一欄填「總計」，"
            "「__QTY__」與「__AMT__」欄**使用 SUM 公式**加總上方所有資料列，其餘欄留空。",
            "幫「__SHEET__」底部加一列總計（第一欄寫「總計」），__QTY__與__AMT__兩欄請用 =SUM(...) 公式，不要填死值。",
        ]), **t)
        ref = fill(_HDR + '''from openpyxl.utils import get_column_letter
last = len(rows) + 1
qi = headers.index("__QTY__") + 1
ai = headers.index("__AMT__") + 1
qL, aL = get_column_letter(qi), get_column_letter(ai)
tr = last + 1
ws.cell(row=tr, column=1, value="總計")
ws.cell(row=tr, column=qi, value=f"=SUM({qL}2:{qL}{last})")
ws.cell(row=tr, column=ai, value=f"=SUM({aL}2:{aL}{last})")
wb.save(OUTPUT_PATH)
''', **t)
        check = {"target_sheets": [t["sheet"]], **_TOL, "formula_cells": [
            {"sheet": t["sheet"], "range": f"{qL}{last + 1}:{qL}{last + 1}", "contains": "SUM("},
            {"sheet": t["sheet"], "range": f"{aL}{last + 1}:{aL}{last + 1}", "contains": "SUM("}]}

    elif variant == "vlookup_price":
        headers, rows, meta = make_table(rng, empty_price_amount=True, consistent_price=True)
        t = schema_tokens(meta["s"])
        ii = meta["idx"][meta["col"]["item"]]
        qi = meta["idx"][meta["col"]["qty"]]
        pi = meta["idx"][meta["col"]["price"]]
        ai = _ai(meta)
        iL, qL = get_column_letter(ii + 1), get_column_letter(qi + 1)
        pL, aL = get_column_letter(pi + 1), get_column_letter(ai + 1)
        last = len(rows) + 1
        price_rows = [[p, meta["price_map"][p]] for p in sorted(meta["price_map"])]
        goal_rows = [list(r) for r in rows]
        for r in goal_rows:
            unit = meta["price_map"][r[ii]]
            r[pi] = unit
            r[ai] = unit * r[qi]
        instruction = fill(rng.choice([
            "「__SHEET__」的「__PRICE__」與「__AMT__」欄是空的。請用 **VLOOKUP 公式**從「__LOOKUP__」帶出"
            "每列的__PRICE__，並用**公式**計算 __AMT__ = __QTY__ × __PRICE__（都不要填死值）。",
            "請幫「__SHEET__」補齊：__PRICE__欄用 VLOOKUP 去「__LOOKUP__」查（依__ITEM__），"
            "__AMT__欄用公式算__QTY__乘__PRICE__。",
        ]), **t)
        ref = fill(_HDR + '''from openpyxl.utils import get_column_letter
last = len(rows) + 1
iL = get_column_letter(headers.index("__ITEM__") + 1)
qL = get_column_letter(headers.index("__QTY__") + 1)
pi = headers.index("__PRICE__") + 1
ai = headers.index("__AMT__") + 1
pL, aL = get_column_letter(pi), get_column_letter(ai)
for r in range(2, last + 1):
    ws.cell(row=r, column=pi, value=f'=VLOOKUP({iL}{r},__LOOKUP__!A:B,2,FALSE)')
    ws.cell(row=r, column=ai, value=f"={qL}{r}*{pL}{r}")
wb.save(OUTPUT_PATH)
''', **t)
        check = {"target_sheets": [t["sheet"], t["lookup"]], **_TOL, "formula_cells": [
            {"sheet": t["sheet"], "range": f"{pL}2:{pL}{last}", "contains": "VLOOKUP"},
            {"sheet": t["sheet"], "range": f"{aL}2:{aL}{last}"}]}
        start = build_wb({t["sheet"]: (headers, rows),
                          t["lookup"]: ([t["item"], t["price"]], price_rows)})
        goal = build_wb({t["sheet"]: (headers, goal_rows),
                         t["lookup"]: ([t["item"], t["price"]], price_rows)})
        spec = TaskSpec(task_id, "formula_write", instruction, check, ref,
                        meta={"variant": variant, "schema": meta["s"]["sheet"]})
        return spec, start, goal

    elif variant == "if_grade":
        headers, rows, meta = make_table(rng, unique_amounts=True)
        t = schema_tokens(meta["s"])
        ai = _ai(meta)
        aL = get_column_letter(ai + 1)
        newL = get_column_letter(len(headers) + 1)
        last = len(rows) + 1
        amts = sorted(r[ai] for r in rows)
        thr = amts[len(amts) // 2]
        goal = build_wb({t["sheet"]: (headers + ["等級"],
                                      [list(r) + ["高" if r[ai] >= thr else "低"] for r in rows])})
        instruction = fill(rng.choice([
            "請在「__SHEET__」最右側新增「等級」欄，**使用 IF 公式**判斷：__AMT__大於或等於 __T__ 填「高」，"
            "否則填「低」（不要用固定文字）。",
            "幫「__SHEET__」加「等級」欄，用 IF 公式：__AMT__ ≥ __T__ 是「高」，其餘是「低」。",
        ]), t=thr, **t)
        ref = fill(_HDR + '''from openpyxl.utils import get_column_letter
aL = get_column_letter(headers.index("__AMT__") + 1)
col = len(headers) + 1
ws.cell(row=1, column=col, value="等級")
for i in range(len(rows)):
    r = i + 2
    ws.cell(row=r, column=col, value=f'=IF({aL}{r}>=__T__,"高","低")')
wb.save(OUTPUT_PATH)
''', t=thr, **t)
        check = {"target_sheets": [t["sheet"]], **_TOL, "formula_cells": [
            {"sheet": t["sheet"], "range": f"{newL}2:{newL}{last}", "contains": "IF("}]}

    elif variant == "round_tax":
        headers, rows, meta = make_table(rng)
        t = schema_tokens(meta["s"])
        ai = _ai(meta)
        aL = get_column_letter(ai + 1)
        newL = get_column_letter(len(headers) + 1)
        last = len(rows) + 1
        # 用整數運算做「四捨五入」以對齊 Excel 語義（Python round() 是銀行家捨入，682.5→682）
        goal = build_wb({t["sheet"]: (headers + ["含稅價"],
                                      [list(r) + [(r[ai] * 105 + 50) // 100] for r in rows])})
        instruction = fill(rng.choice([
            "請在「__SHEET__」最右側新增「含稅價」欄，**使用 ROUND 公式**計算 __AMT__ × 1.05 並四捨五入到整數。",
            "幫「__SHEET__」加「含稅價」欄：用公式 =ROUND(__AMT__*1.05,0)，每列參照自己那列。",
        ]), **t)
        ref = fill(_HDR + '''from openpyxl.utils import get_column_letter
aL = get_column_letter(headers.index("__AMT__") + 1)
col = len(headers) + 1
ws.cell(row=1, column=col, value="含稅價")
for i in range(len(rows)):
    r = i + 2
    ws.cell(row=r, column=col, value=f"=ROUND({aL}{r}*1.05,0)")
wb.save(OUTPUT_PATH)
''', **t)
        check = {"target_sheets": [t["sheet"]], **_TOL, "formula_cells": [
            {"sheet": t["sheet"], "range": f"{newL}2:{newL}{last}", "contains": "ROUND("}]}

    else:  # stats_block：新表放 SUM/AVERAGE/COUNT/MAX 統計
        headers, rows, meta = make_table(rng)
        t = schema_tokens(meta["s"])
        ai = _ai(meta)
        aL = get_column_letter(ai + 1)
        last = len(rows) + 1
        amts = [r[ai] for r in rows]
        stats_rows = [["總金額", sum(amts)], ["平均", sum(amts) / len(amts)],
                      ["筆數", len(amts)], ["最大值", max(amts)]]
        instruction = fill(rng.choice([
            "請新增「統計」工作表：A 欄依序填 總金額、平均、筆數、最大值，B 欄**用公式**"
            "（SUM／AVERAGE／COUNT／MAX）統計「__SHEET__」的「__AMT__」欄。",
            "幫我做一張「統計」工作表：四列分別是 總金額、平均、筆數、最大值，"
            "B 欄請用 SUM、AVERAGE、COUNT、MAX 公式參照「__SHEET__」的__AMT__欄。",
        ]), **t)
        ref = fill(_HDR + '''from openpyxl.utils import get_column_letter
aL = get_column_letter(headers.index("__AMT__") + 1)
last = len(rows) + 1
rng_ = f"'__SHEET__'!{aL}2:{aL}{last}"
out = wb.create_sheet("統計")
out["A1"] = "總金額"; out["B1"] = f"=SUM({rng_})"
out["A2"] = "平均";   out["B2"] = f"=AVERAGE({rng_})"
out["A3"] = "筆數";   out["B3"] = f"=COUNT({rng_})"
out["A4"] = "最大值"; out["B4"] = f"=MAX({rng_})"
wb.save(OUTPUT_PATH)
''', **t)
        check = {"target_sheets": [t["sheet"], "統計"], **_TOL, "formula_cells": [
            {"sheet": "統計", "range": "B1:B4"}]}
        start = build_wb({t["sheet"]: (headers, rows)})
        goal = new_wb()      # 統計表沒有標題列，手工組
        write_table(goal.create_sheet(t["sheet"]), headers, rows)
        gws = goal.create_sheet("統計")
        for row in stats_rows:
            gws.append(row)
        spec = TaskSpec(task_id, "formula_write", instruction, check, ref,
                        meta={"variant": variant, "schema": meta["s"]["sheet"]})
        return spec, start, goal

    spec = TaskSpec(task_id, "formula_write", instruction, check, ref,
                    meta={"variant": variant, "schema": meta["s"]["sheet"]})
    return spec, build_wb({t["sheet"]: (headers, rows)}), goal


V4_FAMILIES = {
    "formula_write": gen_formula_write,
}
