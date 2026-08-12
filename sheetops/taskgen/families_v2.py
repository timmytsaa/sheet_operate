"""v2 難度階梯任務家族（GRPO 的主戰場）。

設計目標：讓 SFT 後的模型成功率落在 30~60%（v1 已飽和於 96%）。
難度來源：長鏈組合（4~6 步）、跨工作表結構操作、欄位結構調整、
格式進階（框線/對齊/合併）、同義詞推斷、多段計算鏈。
"""
from __future__ import annotations

from random import Random

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .base import TaskSpec, build_wb, fill, make_table, new_wb, schema_tokens, write_table
from .families import _HDR, _REWRITE, _ai, _pick_group

# ----------------------------------------------------------------------
# 14. chain_v2 — 長鏈組合（4~6 步）
# ----------------------------------------------------------------------

def gen_chain_v2(rng: Random, task_id: str):
    variant = rng.choice(["six_step", "four_step"])

    if variant == "six_step":
        # 去空白 → 補空值 → 新增金額欄 → 篩選 → 排序 → 總計列
        headers, rows, meta = make_table(rng, with_amount=False,
                                         padded_names=rng.randint(3, 5),
                                         blank_group=rng.randint(2, 3))
        t = schema_tokens(meta["s"])
        ni = meta["idx"][meta["col"]["person"]]
        gi = meta["idx"][meta["col"]["group"]]
        qi, pi = meta["idx"][meta["col"]["qty"]], meta["idx"][meta["col"]["price"]]

        work = [list(r) for r in rows]
        for r in work:
            if isinstance(r[ni], str):
                r[ni] = r[ni].strip()
            if r[gi] is None:
                r[gi] = "未知"
        work = [r + [r[qi] * r[pi]] for r in work]
        amts = sorted(r[-1] for r in work)
        thr = amts[len(amts) // 3]
        kept = sorted([r for r in work if r[-1] >= thr], key=lambda r: r[-1], reverse=True)
        total = [None] * (len(headers) + 1)
        total[0] = "總計"
        total[qi] = sum(r[qi] for r in kept)
        total[-1] = sum(r[-1] for r in kept)
        goal_rows = kept + [total]
        new_headers = headers + [t["amt"]]

        instruction = fill(rng.choice([
            "請對「__SHEET__」工作表依序完成六個步驟：(1) 「__PERSON__」欄去除前後空白；"
            "(2) 「__GROUP__」欄空白填入「未知」；(3) 最右側新增「__AMT__」欄＝__QTY__ × __PRICE__；"
            "(4) 刪除__AMT__小於 __T__ 的資料列；(5) 剩餘資料依「__AMT__」由大到小排序；"
            "(6) 底部新增總計列（第一欄填「總計」，__QTY__與__AMT__欄填總和，其餘留空）。",
            "幫我一次整理「__SHEET__」：__PERSON__名字空白清掉、__GROUP__空值補「未知」、"
            "加一欄「__AMT__」（__QTY__乘__PRICE__）、然後只留__AMT__ ≥ __T__ 的資料並由大到小排好，"
            "最後在底部加總計列（第一欄「總計」，__QTY__、__AMT__兩欄加總，其他空白）。",
        ]), t=thr, **t)
        ref = fill(_HDR + '''ni = headers.index("__PERSON__")
gi = headers.index("__GROUP__")
qi = headers.index("__QTY__")
pi = headers.index("__PRICE__")
work = []
for r in rows:
    r = list(r)
    if isinstance(r[ni], str):
        r[ni] = r[ni].strip()
    if r[gi] is None:
        r[gi] = "未知"
    r.append(r[qi] * r[pi])
    work.append(r)
kept = [r for r in work if r[-1] >= __T__]
kept.sort(key=lambda r: r[-1], reverse=True)
total = [None] * (len(headers) + 1)
total[0] = "總計"
total[qi] = sum(r[qi] for r in kept)
total[-1] = sum(r[-1] for r in kept)
ws.cell(row=1, column=len(headers) + 1, value="__AMT__")
new_rows = kept + [total]
''' + _REWRITE, t=thr, **t)
    else:
        # 刪群組 → 新增折扣價 → 依折扣價排序 → 總計列
        for _ in range(30):
            headers, rows, meta = make_table(rng, group_pool=3, unique_amounts=True)
            gval = _pick_group(rng, rows, meta)
            if gval:
                break
        t = schema_tokens(meta["s"])
        gi = meta["idx"][meta["col"]["group"]]
        pi = meta["idx"][meta["col"]["price"]]
        qi = meta["idx"][meta["col"]["qty"]]

        kept = [list(r) + [r[pi] * 9 // 10] for r in rows if r[gi] != gval]
        kept.sort(key=lambda r: r[-1])
        total = [None] * (len(headers) + 1)
        total[0] = "總計"
        total[qi] = sum(r[qi] for r in kept)
        total[-1] = sum(r[-1] for r in kept)
        goal_rows = kept + [total]
        new_headers = headers + ["折扣價"]

        instruction = fill(rng.choice([
            "請對「__SHEET__」依序完成：(1) 刪除__GROUP__為「__GVAL__」的資料列；"
            "(2) 最右側新增「折扣價」欄＝__PRICE__打九折後無條件捨去；"
            "(3) 依「折扣價」由小到大排序；(4) 底部加總計列（第一欄「總計」，__QTY__與折扣價欄填總和）。",
            "「__SHEET__」先把__GROUP__是「__GVAL__」的列刪掉，加一欄「折扣價」（__PRICE__×0.9 去小數），"
            "照折扣價從低到高排，最後補一列總計（第一欄寫「總計」，__QTY__和折扣價加總）。",
        ]), gval=gval, **t)
        ref = fill(_HDR + '''gi = headers.index("__GROUP__")
pi = headers.index("__PRICE__")
qi = headers.index("__QTY__")
kept = []
for r in rows:
    if r[gi] == "__GVAL__":
        continue
    r = list(r)
    r.append(r[pi] * 9 // 10)
    kept.append(r)
kept.sort(key=lambda r: r[-1])
total = [None] * (len(headers) + 1)
total[0] = "總計"
total[qi] = sum(r[qi] for r in kept)
total[-1] = sum(r[-1] for r in kept)
ws.cell(row=1, column=len(headers) + 1, value="折扣價")
new_rows = kept + [total]
''' + _REWRITE, gval=gval, **t)

    spec = TaskSpec(task_id, "chain_v2", instruction, {"target_sheets": [t["sheet"]]}, ref,
                    meta={"variant": variant, "schema": meta["s"]["sheet"]})
    return spec, build_wb({t["sheet"]: (headers, rows)}), build_wb({t["sheet"]: (new_headers, goal_rows)})


# ----------------------------------------------------------------------
# 15. cross_sheet — 跨工作表結構操作
# ----------------------------------------------------------------------

def gen_cross_sheet(rng: Random, task_id: str):
    variant = rng.choice(["split_by_cat", "pivot", "merge_two"])

    if variant == "split_by_cat":
        headers, rows, meta = make_table(rng)
        t = schema_tokens(meta["s"])
        ci = meta["idx"][meta["col"]["cat"]]
        cats = sorted({r[ci] for r in rows})
        sheets = {t["sheet"]: (headers, rows)}
        goal_sheets = {t["sheet"]: (headers, rows)}
        for cat in cats:
            goal_sheets[cat] = (headers, [r for r in rows if r[ci] == cat])
        instruction = fill(rng.choice([
            "請依「__CAT__」把「__SHEET__」的資料拆分：每個__CAT__建立一個同名的新工作表，"
            "欄位與「__SHEET__」相同，放入該__CAT__的所有資料列（保持原本順序）。「__SHEET__」本身不要動。",
            "幫「__SHEET__」按__CAT__分家：每個__CAT__開一張同名工作表，把屬於它的資料列複製過去"
            "（欄位照舊、順序照舊），原表保留。",
        ]), **t)
        ref = fill(_HDR + '''ci = headers.index("__CAT__")
cats = sorted({r[ci] for r in rows if r[ci] is not None})
for cat in cats:
    out = wb.create_sheet(cat)
    out.append(headers)
    for r in rows:
        if r[ci] == cat:
            out.append(list(r))
wb.save(OUTPUT_PATH)
''', **t)
        check = {"target_sheets": [t["sheet"]] + cats}
        spec = TaskSpec(task_id, "cross_sheet", instruction, check, ref,
                        meta={"variant": variant, "schema": meta["s"]["sheet"]})
        return spec, build_wb(sheets), build_wb(goal_sheets)

    if variant == "pivot":
        headers, rows, meta = make_table(rng, n=rng.randint(16, 22))
        t = schema_tokens(meta["s"])
        gi = meta["idx"][meta["col"]["group"]]
        ci = meta["idx"][meta["col"]["cat"]]
        ai = _ai(meta)
        groups = sorted({r[gi] for r in rows})
        cats = sorted({r[ci] for r in rows})
        grid = {(g, c): 0 for g in groups for c in cats}
        for r in rows:
            grid[(r[gi], r[ci])] += r[ai]
        piv_headers = [t["group"]] + cats
        piv_rows = [[g] + [grid[(g, c)] for c in cats] for g in groups]
        instruction = fill(rng.choice([
            "請新增名為「樞紐」的工作表：第一欄為「__GROUP__」（依名稱排序），"
            "之後每個「__CAT__」一欄（依名稱排序），儲存格填入該__GROUP__×該__CAT__的__AMT__總和"
            "（沒有資料的組合填 0）。「__SHEET__」不要更動。",
            "幫我做一張樞紐表（工作表名「樞紐」）：列是__GROUP__、欄是__CAT__（都按名稱排序），"
            "值是__AMT__加總，空組合補 0，資料來自「__SHEET__」。",
        ]), **t)
        ref = fill(_HDR + '''gi = headers.index("__GROUP__")
ci = headers.index("__CAT__")
ai = headers.index("__AMT__")
groups = sorted({r[gi] for r in rows if r[gi] is not None})
cats = sorted({r[ci] for r in rows if r[ci] is not None})
grid = {}
for r in rows:
    key = (r[gi], r[ci])
    grid[key] = grid.get(key, 0) + r[ai]
out = wb.create_sheet("樞紐")
out.append(["__GROUP__"] + cats)
for g in groups:
    out.append([g] + [grid.get((g, c), 0) for c in cats])
wb.save(OUTPUT_PATH)
''', **t)
        check = {"target_sheets": [t["sheet"], "樞紐"]}
        spec = TaskSpec(task_id, "cross_sheet", instruction, check, ref,
                        meta={"variant": variant, "schema": meta["s"]["sheet"]})
        start = build_wb({t["sheet"]: (headers, rows)})
        goal = build_wb({t["sheet"]: (headers, rows), "樞紐": (piv_headers, piv_rows)})
        return spec, start, goal

    # merge_two：兩張月份表合併 + 標來源
    schema_key = rng.choice(["orders", "expense", "sales", "inventory", "timesheet"])
    headers, rows1, meta = make_table(rng, schema=schema_key, n=rng.randint(8, 12))
    _h2, rows2, _m2 = make_table(rng, schema=schema_key, n=rng.randint(8, 12))
    t = schema_tokens(meta["s"])
    s1, s2 = "1月", "2月"
    merged_rows = [list(r) + [s1] for r in rows1] + [list(r) + [s2] for r in rows2]
    instruction = fill(rng.choice([
        f"活頁簿裡有「{s1}」與「{s2}」兩張同欄位的工作表。請新增名為「合併」的工作表："
        f"欄位為原欄位加上最後一欄「月份」，先放入「{s1}」的全部資料列（月份填「{s1}」），"
        f"再接著放「{s2}」的資料列（月份填「{s2}」）。原兩張表不要更動。",
        f"請把「{s1}」「{s2}」兩張表串成一張新表「合併」：欄位不變、最後加「月份」欄標示資料來源"
        f"（{s1}的列填「{s1}」、{s2}的填「{s2}」），順序是{s1}全部在前、{s2}在後。",
    ]), **t)
    ref = f'''import openpyxl
wb = openpyxl.load_workbook(INPUT_PATH)
out = wb.create_sheet("合併")
first = True
for sname in ["{s1}", "{s2}"]:
    ws = wb[sname]
    headers = [c.value for c in ws[1]]
    if first:
        out.append(headers + ["月份"])
        first = False
    for r in ws.iter_rows(min_row=2, values_only=True):
        out.append(list(r) + [sname])
wb.save(OUTPUT_PATH)
'''
    check = {"target_sheets": [s1, s2, "合併"]}
    spec = TaskSpec(task_id, "cross_sheet", instruction, check, ref,
                    meta={"variant": variant, "schema": meta["s"]["sheet"]})
    start = build_wb({s1: (headers, rows1), s2: (headers, rows2)})
    goal = build_wb({s1: (headers, rows1), s2: (headers, rows2),
                     "合併": (headers + ["月份"], merged_rows)})
    return spec, start, goal


# ----------------------------------------------------------------------
# 16. format_v2 — 進階格式（框線 / 對齊 / 雙色分級 / 標題合併 / 數值格式組合）
# ----------------------------------------------------------------------

def gen_format_v2(rng: Random, task_id: str):
    variant = rng.choice(["grid_border", "align_center", "two_tier", "merge_title", "numfmt_mix"])

    if variant == "merge_title":
        headers, rows, meta = make_table(rng)
        t = schema_tokens(meta["s"])
        last = get_column_letter(len(headers))
        title = f"{t['sheet']}報表"
        # goal：手工建（標題列 + 原表下移一列）
        goal = new_wb()
        gws = goal.create_sheet(t["sheet"])
        write_table(gws, headers, rows)
        gws.insert_rows(1)
        gws.cell(row=1, column=1, value=title)
        gws.merge_cells(f"A1:{last}1")
        c = gws.cell(row=1, column=1)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")
        instruction = fill(rng.choice([
            f"請在「__SHEET__」最上方插入一列作為標題列：A1 填入「{title}」，"
            f"合併 A1:{last}1，文字置中並設為粗體。原有資料整體往下移一列、內容不變。",
            f"幫「__SHEET__」加一個報表標題：頂端插入新列，A1 寫「{title}」，"
            f"把 A1 到 {last}1 合併起來、置中、粗體。",
        ]), **t)
        ref = fill(f'''import openpyxl
from openpyxl.styles import Alignment, Font
wb = openpyxl.load_workbook(INPUT_PATH)
ws = wb["__SHEET__"]
ws.insert_rows(1)
ws.cell(row=1, column=1, value="{title}")
ws.merge_cells("A1:{last}1")
c = ws.cell(row=1, column=1)
c.font = Font(bold=True)
c.alignment = Alignment(horizontal="center")
wb.save(OUTPUT_PATH)
''', **t)
        check = {"target_sheets": [t["sheet"]],
                 "merged": [{"sheet": t["sheet"], "range": f"A1:{last}1"}],
                 "format_checks": [{"sheet": t["sheet"], "range": "A1:A1",
                                    "props": {"bold": True, "align_h": "center"}}]}
        spec = TaskSpec(task_id, "format_v2", instruction, check, ref,
                        meta={"variant": variant, "schema": meta["s"]["sheet"]})
        return spec, build_wb({t["sheet"]: (headers, rows)}), goal

    headers, rows, meta = make_table(rng, unique_amounts=True)
    t = schema_tokens(meta["s"])
    last = get_column_letter(len(headers))
    n_rows = len(rows) + 1
    start = build_wb({t["sheet"]: (headers, rows)},
                     style_header=(variant != "align_center"))
    goal = build_wb({t["sheet"]: (headers, rows)},
                    style_header=(variant != "align_center"))
    gws = goal[t["sheet"]]

    if variant == "grid_border":
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in gws.iter_rows(min_row=1, max_row=n_rows, max_col=len(headers)):
            for cell in row:
                cell.border = border
        instruction = fill(rng.choice([
            f"請為「__SHEET__」工作表的資料範圍 A1:{last}{n_rows} 的每一個儲存格加上細實線框線（上下左右四邊）。",
            f"幫「__SHEET__」的表格（A1 到 {last}{n_rows}）畫上格線：每格四邊都要細框線。",
        ]), **t)
        ref = fill(f'''import openpyxl
from openpyxl.styles import Border, Side
wb = openpyxl.load_workbook(INPUT_PATH)
ws = wb["__SHEET__"]
thin = Side(style="thin")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for row in ws.iter_rows(min_row=1, max_row={n_rows}, max_col={len(headers)}):
    for cell in row:
        cell.border = border
wb.save(OUTPUT_PATH)
''', **t)
        check = {"target_sheets": [t["sheet"]], "format_checks": [
            {"sheet": t["sheet"], "range": f"A1:{last}{n_rows}",
             "props": {"grid_border": True}}]}

    elif variant == "align_center":
        for cell in gws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        instruction = fill(rng.choice([
            "請將「__SHEET__」的標題列（第 1 列）設為粗體並水平置中。",
            "幫「__SHEET__」第一列的欄位標題加粗、文字置中對齊。",
        ]), **t)
        ref = fill('''import openpyxl
from openpyxl.styles import Alignment, Font
wb = openpyxl.load_workbook(INPUT_PATH)
ws = wb["__SHEET__"]
for cell in ws[1]:
    if cell.value is not None:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
wb.save(OUTPUT_PATH)
''', **t)
        check = {"target_sheets": [t["sheet"]], "format_checks": [
            {"sheet": t["sheet"], "range": f"A1:{last}1",
             "props": {"bold": True, "align_h": "center"}}]}

    elif variant == "two_tier":
        ai = _ai(meta)
        amts = sorted(r[ai] for r in rows)
        t_hi = amts[len(amts) * 2 // 3]
        t_lo = amts[len(amts) // 3]
        fcs = []
        for i, r in enumerate(rows):
            if r[ai] >= t_hi:
                for c in range(1, len(headers) + 1):
                    gws.cell(row=i + 2, column=c).fill = PatternFill(
                        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                fcs.append({"sheet": t["sheet"], "range": f"A{i + 2}:{last}{i + 2}",
                            "props": {"fill_rgb": "C6EFCE"}})
            elif r[ai] < t_lo:
                for c in range(1, len(headers) + 1):
                    gws.cell(row=i + 2, column=c).font = Font(color="FF0000")
                fcs.append({"sheet": t["sheet"], "range": f"A{i + 2}:{last}{i + 2}",
                            "props": {"font_rgb": "FF0000"}})
        instruction = fill(rng.choice([
            "請對「__SHEET__」做兩級標示：__AMT__大於或等於 __THI__ 的資料列整列底色設為 #C6EFCE；"
            "__AMT__小於 __TLO__ 的資料列整列字體改為紅色（FF0000）；其餘資料列不動。",
            "幫「__SHEET__」上色分級：__AMT__ ≥ __THI__ 的列鋪綠底（C6EFCE），"
            "__AMT__ < __TLO__ 的列改紅字（FF0000），中間的不用處理。",
        ]), thi=t_hi, tlo=t_lo, **t)
        ref = fill(_HDR + '''from openpyxl.styles import Font, PatternFill
ai = headers.index("__AMT__")
for i, r in enumerate(rows):
    if r[ai] is None:
        continue
    if r[ai] >= __THI__:
        for c in range(1, len(headers) + 1):
            ws.cell(row=i + 2, column=c).fill = PatternFill(
                start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    elif r[ai] < __TLO__:
        for c in range(1, len(headers) + 1):
            ws.cell(row=i + 2, column=c).font = Font(color="FF0000")
wb.save(OUTPUT_PATH)
''', thi=t_hi, tlo=t_lo, **t)
        check = {"target_sheets": [t["sheet"]], "format_checks": fcs}

    else:  # numfmt_mix
        headers, rows, meta = make_table(rng, date_col=True)
        t = schema_tokens(meta["s"])
        last = get_column_letter(len(headers))
        n_rows = len(rows) + 1
        start = build_wb({t["sheet"]: (headers, rows)})
        goal = build_wb({t["sheet"]: (headers, rows)})
        gws = goal[t["sheet"]]
        di = meta["idx"][meta["col"]["date"]] + 1
        ai = _ai(meta) + 1
        fcs = []
        for ci, fmt in ((di, "yyyy/mm/dd"), (ai, "#,##0")):
            colL = get_column_letter(ci)
            for r in range(2, n_rows + 1):
                gws.cell(row=r, column=ci).number_format = fmt
            fcs.append({"sheet": t["sheet"], "range": f"{colL}2:{colL}{n_rows}",
                        "props": {"number_format": fmt}})
        instruction = fill(rng.choice([
            "請調整「__SHEET__」兩個欄位的數值格式（僅資料列）：「__DATE_H__」欄改為 yyyy/mm/dd，"
            "「__AMT__」欄改為千分位 #,##0。",
            "幫「__SHEET__」改格式：__DATE_H__欄顯示成 yyyy/mm/dd、__AMT__欄用千分位（#,##0），標題列不用動。",
        ]), **t)
        ref = fill(_HDR + '''di = headers.index("__DATE_H__") + 1
ai = headers.index("__AMT__") + 1
for r in range(2, ws.max_row + 1):
    ws.cell(row=r, column=di).number_format = "yyyy/mm/dd"
    ws.cell(row=r, column=ai).number_format = "#,##0"
wb.save(OUTPUT_PATH)
''', **t)
        check = {"target_sheets": [t["sheet"]], "format_checks": fcs}

    spec = TaskSpec(task_id, "format_v2", instruction, check, ref,
                    meta={"variant": variant, "schema": meta["s"]["sheet"]})
    return spec, start, goal


# ----------------------------------------------------------------------
# 17. column_ops — 欄位結構操作
# ----------------------------------------------------------------------

def gen_column_ops(rng: Random, task_id: str):
    variant = rng.choice(["insert_after", "delete_col", "move_col", "reorder"])
    headers, rows, meta = make_table(rng)
    t = schema_tokens(meta["s"])

    if variant == "insert_after":
        qi = meta["idx"][meta["col"]["qty"]]
        new_headers = headers[:qi + 1] + ["備註"] + headers[qi + 1:]
        goal_rows = [r[:qi + 1] + [None] + r[qi + 1:] for r in rows]
        instruction = fill(rng.choice([
            "請在「__SHEET__」的「__QTY__」欄右邊插入一個新欄位，標題填「備註」，資料列留空，"
            "其他欄位與內容位置順勢右移。",
            "幫「__SHEET__」加一欄「備註」（空白欄），位置在「__QTY__」欄的正右邊。",
        ]), **t)
        ref = fill('''import openpyxl
wb = openpyxl.load_workbook(INPUT_PATH)
ws = wb["__SHEET__"]
headers = [c.value for c in ws[1]]
idx = headers.index("__QTY__") + 1
ws.insert_cols(idx + 1)
ws.cell(row=1, column=idx + 1, value="備註")
wb.save(OUTPUT_PATH)
''', **t)
    elif variant == "delete_col":
        gi = meta["idx"][meta["col"]["group"]]
        new_headers = headers[:gi] + headers[gi + 1:]
        goal_rows = [r[:gi] + r[gi + 1:] for r in rows]
        instruction = fill(rng.choice([
            "請刪除「__SHEET__」的「__GROUP__」整欄，其餘欄位順勢左移、內容不變。",
            "「__SHEET__」用不到「__GROUP__」欄了，請整欄移除。",
        ]), **t)
        ref = fill('''import openpyxl
wb = openpyxl.load_workbook(INPUT_PATH)
ws = wb["__SHEET__"]
headers = [c.value for c in ws[1]]
idx = headers.index("__GROUP__") + 1
ws.delete_cols(idx)
wb.save(OUTPUT_PATH)
''', **t)
    elif variant == "move_col":
        ai = _ai(meta)
        pi = meta["idx"][meta["col"]["person"]]
        order = [h for h in headers if h != t["amt"]]
        order.insert(order.index(t["person"]) + 1, t["amt"])
        imap = [headers.index(h) for h in order]
        new_headers = order
        goal_rows = [[r[i] for i in imap] for r in rows]
        instruction = fill(rng.choice([
            "請把「__SHEET__」的「__AMT__」欄移動到「__PERSON__」欄的右邊，其餘欄位相對順序不變。",
            "調整「__SHEET__」欄位位置：「__AMT__」整欄搬到「__PERSON__」旁邊（右側），別的欄不動相對順序。",
        ]), **t)
        ref = fill(_HDR + '''order = [h for h in headers if h != "__AMT__"]
order.insert(order.index("__PERSON__") + 1, "__AMT__")
imap = [headers.index(h) for h in order]
ws.delete_rows(1, ws.max_row)
ws.append(order)
for r in rows:
    ws.append([r[i] for i in imap])
wb.save(OUTPUT_PATH)
''', **t)
    else:  # reorder
        perm = list(range(len(headers)))
        rng.shuffle(perm)
        new_headers = [headers[i] for i in perm]
        goal_rows = [[r[i] for i in perm] for r in rows]
        order_str = "、".join(new_headers)
        instruction = fill(
            f"請將「__SHEET__」的欄位重新排列成以下順序（由左到右）：{order_str}。"
            "各欄內容跟著欄位一起移動，資料列順序不變。", **t)
        ref = fill(_HDR + f'''order = {new_headers!r}
imap = [headers.index(h) for h in order]
ws.delete_rows(1, ws.max_row)
ws.append(order)
for r in rows:
    ws.append([r[i] for i in imap])
wb.save(OUTPUT_PATH)
''', **t)

    spec = TaskSpec(task_id, "column_ops", instruction, {"target_sheets": [t["sheet"]]}, ref,
                    meta={"variant": variant, "schema": meta["s"]["sheet"]})
    return spec, build_wb({t["sheet"]: (headers, rows)}), build_wb({t["sheet"]: (new_headers, goal_rows)})


# ----------------------------------------------------------------------
# 18. semantic_map — 同義詞指令（欄位名不直說，考語意對映）
# ----------------------------------------------------------------------

_SYNONYMS = {
    "orders":    {"amount": "營收", "person": "買家", "group": "據點"},
    "expense":   {"amount": "開銷", "person": "同仁", "group": "單位"},
    "inventory": {"amount": "庫存價值", "person": "管理人", "group": "存放位置"},
    "timesheet": {"amount": "成本", "person": "投入人員", "group": "案子"},
    "sales":     {"amount": "業績", "person": "銷售人員", "group": "通路"},
}


def gen_semantic_map(rng: Random, task_id: str):
    variant = rng.choice(["filter_syn", "sort_syn", "top_syn"])
    headers, rows, meta = make_table(rng, unique_amounts=True)
    t = schema_tokens(meta["s"])
    syn = _SYNONYMS.get(meta["key"],
                        {"amount": "價值", "person": "人員", "group": "分組"})
    ai = _ai(meta)

    if variant == "filter_syn":
        amts = sorted(r[ai] for r in rows)
        thr = amts[len(amts) // 2]
        goal_sheets = {t["sheet"]: (headers, [r for r in rows if r[ai] >= thr])}
        check = {"target_sheets": [t["sheet"]]}
        instruction = f"「{t['sheet']}」裡{syn['amount']}沒達到 {thr} 的那些資料都不要了，請整列刪掉。"
        ref = fill(_HDR + '''ai = headers.index("__AMT__")
new_rows = [r for r in rows if r[ai] is not None and r[ai] >= __T__]
''' + _REWRITE, t=thr, **t)
    elif variant == "sort_syn":
        goal_sheets = {t["sheet"]: (headers, sorted(rows, key=lambda r: r[ai], reverse=True))}
        check = {"target_sheets": [t["sheet"]]}
        instruction = f"幫我把「{t['sheet']}」照{syn['amount']}高低排一下，{syn['amount']}最高的放最上面。"
        ref = fill(_HDR + '''ai = headers.index("__AMT__")
new_rows = sorted(rows, key=lambda r: r[ai], reverse=True)
''' + _REWRITE, **t)
    else:  # top_syn
        n_top = rng.randint(3, 5)
        top = sorted(rows, key=lambda r: r[ai], reverse=True)[:n_top]
        sheet2 = f"TOP{n_top}"
        goal_sheets = {t["sheet"]: (headers, rows), sheet2: (headers, top)}
        check = {"target_sheets": [t["sheet"], sheet2]}
        instruction = (f"想看{syn['amount']}最好的前 {n_top} 筆：請開一張「{sheet2}」工作表，"
                       f"把「{t['sheet']}」裡{syn['amount']}最高的 {n_top} 筆由高到低放進去，欄位照舊，原表不動。")
        ref = fill(_HDR + f'''ai = headers.index("__AMT__")
top = sorted(rows, key=lambda r: r[ai], reverse=True)[:{n_top}]
out = wb.create_sheet("{sheet2}")
out.append(headers)
for r in top:
    out.append(list(r))
wb.save(OUTPUT_PATH)
''', **t)

    spec = TaskSpec(task_id, "semantic_map", instruction, check, ref,
                    meta={"variant": variant, "schema": meta["s"]["sheet"], "syn": syn["amount"]})
    return spec, build_wb({t["sheet"]: (headers, rows)}), build_wb(goal_sheets)


# ----------------------------------------------------------------------
# 19. calc_chain — 多段計算鏈
# ----------------------------------------------------------------------

def gen_calc_chain(rng: Random, task_id: str):
    variant = rng.choice(["tier_label", "cum_sum", "pct_share"])
    headers, rows, meta = make_table(rng, unique_amounts=True)
    t = schema_tokens(meta["s"])
    ai = _ai(meta)

    if variant == "tier_label":
        amts = sorted(r[ai] for r in rows)
        t_hi = amts[len(amts) * 2 // 3]
        t_lo = amts[len(amts) // 3]
        goal_rows = [r + [("A" if r[ai] >= t_hi else ("B" if r[ai] >= t_lo else "C"))] for r in rows]
        new_headers = headers + ["等級"]
        instruction = fill(rng.choice([
            "請在「__SHEET__」最右側新增「等級」欄：__AMT__ ≥ __THI__ 填「A」，"
            "__AMT__ ≥ __TLO__（但小於 __THI__）填「B」，其餘填「C」。",
            "幫「__SHEET__」加一欄「等級」做分級：__THI__ 以上是 A、__TLO__ 到 __THI__ 之間是 B、"
            "不到 __TLO__ 的是 C。",
        ]), thi=t_hi, tlo=t_lo, **t)
        ref = fill(_HDR + '''ai = headers.index("__AMT__")
col = len(headers) + 1
ws.cell(row=1, column=col, value="等級")
for i, r in enumerate(rows):
    if r[ai] >= __THI__:
        label = "A"
    elif r[ai] >= __TLO__:
        label = "B"
    else:
        label = "C"
    ws.cell(row=i + 2, column=col, value=label)
wb.save(OUTPUT_PATH)
''', thi=t_hi, tlo=t_lo, **t)
    elif variant == "cum_sum":
        ordered = sorted(rows, key=lambda r: r[ai], reverse=True)
        cum, goal_rows = 0, []
        for r in ordered:
            cum += r[ai]
            goal_rows.append(list(r) + [cum])
        new_headers = headers + [f"累計{t['amt']}"]
        instruction = fill(
            "請對「__SHEET__」依序完成：(1) 資料列依「__AMT__」由大到小排序；"
            "(2) 最右側新增「累計__AMT__」欄，值為排序後由上而下的__AMT__累計總和。", **t)
        ref = fill(_HDR + '''ai = headers.index("__AMT__")
ordered = sorted(rows, key=lambda r: r[ai], reverse=True)
cum = 0
new_rows = []
for r in ordered:
    cum += r[ai]
    new_rows.append(list(r) + [cum])
ws.cell(row=1, column=len(headers) + 1, value="累計__AMT__")
''' + _REWRITE, **t)
    else:  # pct_share
        total = sum(r[ai] for r in rows)
        goal_rows = [r + [round(r[ai] / total, 4)] for r in rows]
        new_headers = headers + ["占比"]
        instruction = fill(
            "請在「__SHEET__」最右側新增「占比」欄＝該列__AMT__ ÷ 全部__AMT__總和，"
            "四捨五入到小數第 4 位（填數值，不用百分比格式）。", **t)
        ref = fill(_HDR + '''ai = headers.index("__AMT__")
total = sum(r[ai] for r in rows if r[ai] is not None)
col = len(headers) + 1
ws.cell(row=1, column=col, value="占比")
for i, r in enumerate(rows):
    ws.cell(row=i + 2, column=col, value=round(r[ai] / total, 4))
wb.save(OUTPUT_PATH)
''', **t)

    spec = TaskSpec(task_id, "calc_chain", instruction, {"target_sheets": [t["sheet"]]}, ref,
                    meta={"variant": variant, "schema": meta["s"]["sheet"]})
    return spec, build_wb({t["sheet"]: (headers, rows)}), build_wb({t["sheet"]: (new_headers, goal_rows)})


V2_FAMILIES = {
    "chain_v2": gen_chain_v2,
    "cross_sheet": gen_cross_sheet,
    "format_v2": gen_format_v2,
    "column_ops": gen_column_ops,
    "semantic_map": gen_semantic_map,
    "calc_chain": gen_calc_chain,
}
