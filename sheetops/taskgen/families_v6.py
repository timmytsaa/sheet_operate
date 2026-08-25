"""v6 任務家族：欄位定位穩健性（column resolution robustness）。

來源：真實 BOM 檔實測（同一產品的客戶版與內部版兩份）。同一個病根在三個不同檔案重現——
模型用「數欄位」定位（硬編 row[11]、column=14），而不是「查表頭名稱」定位：

  1. Parrot「All」表：Note / Usage / Note (OOB) 三個 note-like 欄 → 抓錯欄
  2. 客戶版 BOM：E 欄與 R 欄都叫「M/S」→ 選到空白欄 → 輸出空表（靜默錯誤）
  3. 內部版三張分部門表合併：欄位順序與數量都不同 → row[15] 越界當掉

注意：統計分析本身模型已經會了（實測 Category 分組計數＋Qty 合計＋降冪排序全對），
所以 v6 不補統計，補的是「取到正確的那一欄」。

四個變體
--------
dup_header       同表有重名欄，需靠第 1 列的合併群組標題消歧（← 客戶版的兩個 M/S）
misaligned_merge 多表合併，各表欄位順序/數量不同，只能按表頭名稱對齊（← 內部版三表合併）
two_tier_header  表頭跨兩列＋合併群組標題，第 2 列左半還混著資料值（← 客戶版的版型）
pair_group       同鍵多列的主/替配對，需整組取出或加總時排除替代料（← BOM M/S 語意）

設計原則
--------
1. **必須有鑑別度**：每個產生器都用 _assert_discriminates() 檢查「用錯欄會得到不同答案」。
   若硬編索引剛好也對，這題就沒有教學價值，直接重抽。
2. 參考解法一律示範正確技術：先讀表頭列建 name → index 對照，再取值；
   跨表時每張表各建一份對照，缺欄補 None。
3. 新工作表由模型自行命名 → 驗證走 check["new_sheet"]（只比內容不比名字）。
"""
from __future__ import annotations

from random import Random

from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .. import zh_data
from .base import TaskSpec, new_wb, write_table

# ----------------------------------------------------------------------
# 詞庫（BOM／對照表情境，刻意與 v1~v5 的商業表皮不同）
# ----------------------------------------------------------------------

_PART_NAMES = [
    "上蓋", "下蓋", "中框", "導光柱", "散熱片", "彈簧", "螺絲 M3x6", "螺絲 M2x4",
    "泡棉墊", "銅柱", "排線", "天線", "電源線", "絕緣片", "麥拉片", "腳墊",
    "SIM 卡托", "防塵蓋", "軸承", "卡榫", "遮光罩", "固定夾", "束帶", "標籤",
]
_PART_CATS = ["塑膠件", "金屬件", "五金", "線材", "包材", "電子件"]
_UNITS = ["EA", "PCS", "SET", "M"]

# 區塊標題（第 1 列的合併群組標題）
_BLOCK_PAIRS = [
    ("原廠", "本廠"), ("客戶提供", "我方確認"), ("舊版", "新版"),
    ("A 廠", "B 廠"), ("設計值", "實測值"),
]
_BASE_TITLES = ["基本資料", "料件資訊", None]

# 主料／替代料的標記詞彙
_MS_PAIRS = [("M", "S"), ("主", "替"), ("主料", "替代料"), ("正式", "備選")]

_BOM_SHEET_NAMES = ["BOM", "料表", "零件清單", "用料明細", "主件表", "組立清單"]


def _part_no(rng: Random, i: int) -> str:
    return f"{rng.choice(['30', '31', '32', '39'])}{rng.randint(1, 9)}-{i:05d}"


def _pick_distinct(rng: Random, pool: list, k: int) -> list:
    p = pool[:]
    rng.shuffle(p)
    return p[:k]


def _bold_row(ws, row: int, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        ws.cell(row=row, column=c).font = Font(bold=True)


# ======================================================================
# 變體 1：dup_header —— 同表重名欄，靠合併群組標題消歧
# ======================================================================

def _build_dup_header(rng: Random):
    n = rng.randint(14, 22)
    sentinel = rng.choice(zh_data.PASS_SENTINELS)
    blocks = list(rng.choice(_BLOCK_PAIRS))
    rng.shuffle(blocks)                      # 目標區塊不固定在左邊或右邊
    base_title = rng.choice(_BASE_TITLES)
    status_col = rng.choice(["狀態", "檢驗結果", "判定", "M/S"])
    sub_cols = ["料號", "數量", status_col]  # 這三個欄名在兩個區塊各出現一次

    headers = ["項次", "品名"] + sub_cols + sub_cols
    names = _pick_distinct(rng, _PART_NAMES, n)
    defects = _pick_distinct(rng, zh_data.DEFECT_TEXTS, len(zh_data.DEFECT_TEXTS))

    rows, di = [], 0
    for i in range(n):
        row = [i + 1, names[i]]
        for _ in range(2):                   # 兩個區塊各一組 料號/數量/狀態
            bad = rng.random() < 0.35
            if bad and di < len(defects):
                status = defects[di]
                di += 1
            else:
                status = sentinel
            row += [_part_no(rng, 1000 + rng.randint(0, 8999)), rng.randint(1, 6), status]
        rows.append(row)

    # 兩個區塊的狀態欄分別在第 4 欄(idx 4)與第 7 欄(idx 7)
    st_idx = {blocks[0]: 4, blocks[1]: 7}
    target_block = rng.choice(blocks)
    ti = st_idx[target_block]
    other = st_idx[blocks[0] if blocks[1] == target_block else blocks[1]]

    hit = [r for r in rows if str(r[ti]).strip() != sentinel]
    miss = [r for r in rows if str(r[other]).strip() != sentinel]
    _assert_discriminates(hit, miss, "dup_header：兩區塊的不合格列相同，用錯欄也會對")

    return {
        "n_cols": len(headers), "headers": headers, "rows": rows,
        "blocks": blocks, "base_title": base_title, "sentinel": sentinel,
        "status_col": status_col, "target_block": target_block,
        "expected": hit, "sheet": rng.choice(_BOM_SHEET_NAMES),
    }


def _assert_discriminates(expected: list, naive: list, msg: str) -> None:
    """用錯欄／用錯方法必須得到不同答案，否則這題沒有鑑別度。"""
    assert len(expected) >= 3, f"{msg}（正解列數太少：{len(expected)}）"
    assert [tuple(map(str, r)) for r in expected] != [tuple(map(str, r)) for r in naive], msg


def _tier(rng: Random) -> str:
    """指令揭露層級——初版把版型判讀直接寫進指令，等於送答案。

    真實使用紀錄（logs/usage_log.jsonl）指令長度中位數 35 字，v6 初版卻是 95 字，
    因為我寫了「第 1 列是區塊標題、第 2 列才是欄位名稱，資料從第 3 列開始」這種說明。
    真實使用者只會說「本廠那邊有問題的挑出來」。

      full  說明版型（表頭在第幾列、群組怎麼對應）
      mid   只點出欄名重複／表頭不在第一列，不說怎麼解
      terse 完全不提版型

    砍掉的只有版型提示；輸出欄位一律保留（驗證需要唯一答案）。
    答案仍然唯一——合併儲存格的群組標題在資料裡看得見。
    """
    r = rng.random()
    return "full" if r < 0.34 else ("mid" if r < 0.67 else "terse")


def _compose(rng: Random, tiers: dict[str, list[str]]) -> str:
    return rng.choice(tiers[_tier(rng)])


def _write_dup_header(c: dict, with_result: bool):
    wb = new_wb()
    ws = wb.create_sheet(c["sheet"])
    # 第 1 列：區塊標題（合併儲存格，只有最左格有值——真實檔就長這樣）
    if c["base_title"]:
        ws.cell(row=1, column=1, value=c["base_title"])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    ws.cell(row=1, column=3, value=c["blocks"][0])
    ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=5)
    ws.cell(row=1, column=6, value=c["blocks"][1])
    ws.merge_cells(start_row=1, start_column=6, end_row=1, end_column=8)
    _bold_row(ws, 1, c["n_cols"])
    # 第 2 列：欄名（料號/數量/狀態 各出現兩次）
    for j, h in enumerate(c["headers"]):
        ws.cell(row=2, column=j + 1, value=h)
    _bold_row(ws, 2, c["n_cols"])
    for i, r in enumerate(c["rows"]):
        for j, v in enumerate(r):
            ws.cell(row=3 + i, column=j + 1, value=v)
    for ci in range(1, c["n_cols"] + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 14
    if with_result:
        out = wb.create_sheet(f"{c['target_block']}_異常")
        out.append(c["headers"])
        for r in c["expected"]:
            out.append(list(r))
    return wb


def gen_dup_header(rng: Random, task_id: str):
    for _ in range(40):
        try:
            c = _build_dup_header(rng)
            break
        except AssertionError:
            continue
    else:
        raise RuntimeError("dup_header：40 次都建不出有鑑別度的資料")

    instruction = _compose(rng, {
        "full": [
            f"「{c['sheet']}」表的第 1 列是區塊標題、第 2 列才是欄位名稱，資料從第 3 列開始。"
            f"表裡「{c['status_col']}」欄出現兩次，請取「{c['target_block']}」區塊底下的那一欄，"
            f"把值不是「{c['sentinel']}」的資料列挑出來放到新工作表，欄位與第 2 列表頭相同。",

            f"請在「{c['sheet']}」表裡找出「{c['target_block']}」這個區塊的「{c['status_col']}」欄"
            f"（第 1 列的合併標題是區塊名，第 2 列是欄名，兩個區塊的欄名一樣），"
            f"把該欄不等於「{c['sentinel']}」的整列複製到新工作表，表頭沿用第 2 列。",
        ],
        "mid": [
            f"「{c['sheet']}」是左右兩個區塊的對照表，欄名有重複。"
            f"只看「{c['target_block']}」區的「{c['status_col']}」，"
            f"不是「{c['sentinel']}」的列挑出來做成新工作表，表頭沿用原表。",

            f"請挑出「{c['sheet']}」表裡「{c['target_block']}」的「{c['status_col']}」"
            f"不是「{c['sentinel']}」的資料列（同名欄位不只一個，別取錯），"
            f"放到新工作表，欄位與原表相同。",
        ],
        "terse": [
            f"「{c['sheet']}」表裡「{c['target_block']}」的「{c['status_col']}」"
            f"不是「{c['sentinel']}」的挑出來做成新表，欄位與原表相同。",

            f"把「{c['target_block']}」那邊「{c['status_col']}」有問題的"
            f"（不是「{c['sentinel']}」）整列挑到新工作表，欄位照原表。",

            f"「{c['sheet']}」的「{c['target_block']}」區，"
            f"「{c['status_col']}」非「{c['sentinel']}」的做成新工作表，表頭同原表。",
        ],
    })

    start_wb, goal_wb = _write_dup_header(c, False), _write_dup_header(c, True)
    known = [c["sheet"]]

    ref = f'''# 推斷：資料表={c["sheet"]}（群組標題第1列、表頭第2列、資料第3列起）｜欄位={c["status_col"]}（限「{c["target_block"]}」區塊）｜篩選=值≠"{c["sentinel"]}"｜輸出=新工作表
import openpyxl
wb = openpyxl.load_workbook(INPUT_PATH)
ws = wb["{c["sheet"]}"]

GROUP_ROW, HEADER_ROW = 1, 2
# 合併儲存格只有最左格有值 → 往右延續，得到每一欄所屬的區塊
groups, cur = [], None
for col in range(1, ws.max_column + 1):
    v = ws.cell(row=GROUP_ROW, column=col).value
    if v is not None and str(v).strip() != "":
        cur = str(v).strip()
    groups.append(cur)
headers = [ws.cell(row=HEADER_ROW, column=col).value for col in range(1, ws.max_column + 1)]

def col_index(group, name):
    """欄名重複時，用區塊標題消歧——不可以硬編欄號。"""
    for i, (g, h) in enumerate(zip(groups, headers)):
        if g == group and h == name:
            return i
    raise ValueError("找不到欄位：" + str(group) + "/" + str(name))

st_i = col_index("{c["target_block"]}", "{c["status_col"]}")

out = wb.create_sheet("{c["target_block"]}_異常")
out.append(headers)
for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
    if row[0] is None:
        continue
    v = row[st_i]
    if v is None or str(v).strip() == "{c["sentinel"]}":
        continue
    out.append(list(row))
wb.save(OUTPUT_PATH)
'''

    check = {"target_sheets": known,
             "new_sheet": {"known_sheets": known,
                           "headers": c["headers"],
                           "rows": [list(r) for r in c["expected"]]}}
    spec = TaskSpec(task_id, "dup_header", instruction, check, ref,
                    meta={"variant": "dup_header", "block": c["target_block"],
                          "status_col": c["status_col"], "sentinel": c["sentinel"],
                          "n_expected": len(c["expected"])})
    return spec, start_wb, goal_wb


# ======================================================================
# 變體 2：misaligned_merge —— 多表欄位順序/數量不同的合併
# ======================================================================

_MERGE_BASE = ["項次", "料號", "品名", "數量", "單位", "類別"]


def _build_misaligned_merge(rng: Random):
    sheet_names = _pick_distinct(rng, ["機構件", "電子件", "包材", "ME", "EE", "附件", "選配件"], 3)
    target_cols = _MERGE_BASE[:]
    rng.shuffle(target_cols)
    target_cols = ["項次"] + [c for c in target_cols if c != "項次"]   # 項次固定第一欄

    layouts = [target_cols]
    # 第 2 張：換順序＋少一欄
    l2 = [c for c in target_cols if c != rng.choice(["單位", "類別"])]
    rng.shuffle(l2)
    layouts.append(["項次"] + [c for c in l2 if c != "項次"])
    # 第 3 張：換順序＋多一欄（多出來的欄不在目標表頭內，合併時應丟棄）
    l3 = target_cols[:]
    rng.shuffle(l3)
    l3 = ["項次"] + [c for c in l3 if c != "項次"]
    extra = rng.choice(["備註", "Sub. rank", "Dri"])
    l3.insert(rng.randint(1, len(l3)), extra)
    layouts.append(l3)

    sheets, expected = [], []
    for si, (name, cols) in enumerate(zip(sheet_names, layouts)):
        n = rng.randint(4, 8)
        parts = _pick_distinct(rng, _PART_NAMES, n)
        rows = []
        for i in range(n):
            rec = {"項次": i + 1, "料號": _part_no(rng, 1000 + rng.randint(0, 8999)),
                   "品名": parts[i], "數量": rng.randint(1, 12),
                   "單位": rng.choice(_UNITS), "類別": rng.choice(_PART_CATS),
                   "備註": "", "Sub. rank": "", "Dri": rng.choice(["ME", "EE"])}
            rows.append([rec.get(c) for c in cols])
            expected.append([name] + [rec.get(c) if c in cols else None for c in target_cols])
        sheets.append({"name": name, "cols": cols, "rows": rows})

    # 鑑別度：至少兩張表的欄位順序與目標不同，且欄位集合不完全相同
    diff_order = sum(1 for s in sheets if s["cols"] != target_cols)
    assert diff_order >= 2, "misaligned_merge：欄位順序都一樣，硬編索引也會對"
    assert any(set(s["cols"]) != set(target_cols) for s in sheets), \
        "misaligned_merge：欄位集合完全相同，缺欄補空的邏輯考不到"

    return {"sheets": sheets, "target_cols": target_cols,
            "target_sheet": sheet_names[0], "expected": expected}


def _write_misaligned_merge(c: dict, with_result: bool):
    wb = new_wb()
    for s in c["sheets"]:
        write_table(wb.create_sheet(s["name"]), s["cols"], s["rows"])
    if with_result:
        out = wb.create_sheet("合併清單")
        out.append(["來源表"] + c["target_cols"])
        for r in c["expected"]:
            out.append(list(r))
    return wb


def gen_misaligned_merge(rng: Random, task_id: str):
    for _ in range(40):
        try:
            c = _build_misaligned_merge(rng)
            break
        except AssertionError:
            continue
    else:
        raise RuntimeError("misaligned_merge：40 次都建不出有鑑別度的資料")

    names = "、".join(f"「{s['name']}」" for s in c["sheets"])
    instruction = _compose(rng, {
        "full": [
            f"這三張表 {names} 欄位順序不一樣、有的多一欄有的少一欄。"
            f"請依「{c['target_sheet']}」表的欄位順序合併成一張新工作表，"
            f"第一欄加「來源表」寫上來源工作表名稱，缺的欄位留白，多的欄位捨棄。",

            f"請做一張合併總表（新工作表）：依序放 {names} 的資料，"
            f"欄位為「來源表」＋「{c['target_sheet']}」表的表頭；各表欄位順序不同，"
            f"要按欄名對齊，沒有的欄位留空。",
        ],
        "mid": [
            f"把 {names} 三張表合併成一張新工作表，欄位順序照「{c['target_sheet']}」表，"
            f"最前面加一欄「來源表」記錄資料來自哪張表（三張表的欄位不完全一致）。",

            f"請把 {names} 合併成一張新工作表，"
            f"表頭用「來源表」＋「{c['target_sheet']}」的欄位，注意各表欄位對不齊。",
        ],
        "terse": [
            f"把 {names} 合併成一張新工作表，欄位照「{c['target_sheet']}」表，"
            f"最前面加一欄「來源表」。",

            f"{names} 三張表併成一張新表，欄位順序跟「{c['target_sheet']}」一樣，"
            f"第一欄放「來源表」。",

            f"請將 {names} 整合到新工作表，"
            f"第一欄「來源表」，其餘欄位依「{c['target_sheet']}」表。",
        ],
    })

    start_wb, goal_wb = _write_misaligned_merge(c, False), _write_misaligned_merge(c, True)
    known = [s["name"] for s in c["sheets"]]
    sheet_list = ", ".join(f'"{s["name"]}"' for s in c["sheets"])

    ref = f'''# 推斷：資料表={known}（表頭第1列）｜欄位=按表頭名稱對齊（各表順序不同）｜輸出=新工作表「合併清單」
import openpyxl
wb = openpyxl.load_workbook(INPUT_PATH)

SHEETS = [{sheet_list}]
TARGET = [cell.value for cell in wb["{c["target_sheet"]}"][1]]

out = wb.create_sheet("合併清單")
out.append(["來源表"] + TARGET)

for name in SHEETS:
    ws = wb[name]
    headers = [cell.value for cell in ws[1]]
    # 每張表各建一份「欄名 → 索引」對照；絕不沿用別張表的欄號
    pos = {{h: i for i, h in enumerate(headers)}}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[pos["項次"]] is None:
            continue
        out.append([name] + [row[pos[h]] if h in pos else None for h in TARGET])

wb.save(OUTPUT_PATH)
'''

    check = {"target_sheets": known,
             "new_sheet": {"known_sheets": known,
                           "headers": ["來源表"] + c["target_cols"],
                           "rows": [list(r) for r in c["expected"]]}}
    spec = TaskSpec(task_id, "misaligned_merge", instruction, check, ref,
                    meta={"variant": "misaligned_merge", "sheets": known,
                          "target_cols": c["target_cols"], "n_expected": len(c["expected"])})
    return spec, start_wb, goal_wb


# ======================================================================
# 變體 3：two_tier_header —— 表頭跨兩列，第 2 列左半還混著資料值
# ======================================================================

def _build_two_tier(rng: Random):
    n = rng.randint(14, 22)
    sentinel = rng.choice(zh_data.PASS_SENTINELS)
    group_title = rng.choice(["檢驗項目", "相關單位簽名確認", "品保確認", "各站結果"])
    items = _pick_distinct(rng, ["外觀", "尺寸", "功能", "電性", "耐候", "包裝"], 3)
    left = ["項次", "料號", "品名", "數量"]
    names = _pick_distinct(rng, _PART_NAMES, n)
    defects = _pick_distinct(rng, zh_data.DEFECT_TEXTS, len(zh_data.DEFECT_TEXTS))

    rows, di = [], 0
    for i in range(n):
        row = [i + 1, _part_no(rng, 1000 + rng.randint(0, 8999)), names[i], rng.randint(1, 9)]
        for _ in items:
            if rng.random() < 0.3 and di < len(defects):
                row.append(defects[di])
                di += 1
            else:
                row.append(sentinel)
        rows.append(row)

    target_item = rng.choice(items)
    ti = len(left) + items.index(target_item)
    hit = [r for r in rows if str(r[ti]).strip() != sentinel]
    for other in items:
        if other == target_item:
            continue
        oi = len(left) + items.index(other)
        _assert_discriminates(hit, [r for r in rows if str(r[oi]).strip() != sentinel],
                              "two_tier_header：不同檢驗項目的不合格列相同，用錯欄也會對")

    # 第 2 列左半塞的是「母階資訊」——真實檔案就是這樣，讓天真的表頭偵測踩空
    parent_row = [None, _part_no(rng, 20000 + rng.randint(0, 999)),
                  rng.choice(["SET ASSY", "母件組立", "PACKING ASSY"]), 1]

    out_cols = ["項次", "料號", "品名", target_item]
    expected = [[r[0], r[1], r[2], r[ti]] for r in hit]
    return {"n": n, "left": left, "items": items, "group_title": group_title,
            "rows": rows, "parent_row": parent_row, "sentinel": sentinel,
            "target_item": target_item, "out_cols": out_cols, "expected": expected,
            "sheet": rng.choice(_BOM_SHEET_NAMES)}


def _write_two_tier(c: dict, with_result: bool):
    wb = new_wb()
    ws = wb.create_sheet(c["sheet"])
    n_left, n_items = len(c["left"]), len(c["items"])
    total = n_left + n_items
    # 第 1 列：左半是真正的欄名，右半是跨欄合併的群組標題
    for j, h in enumerate(c["left"]):
        ws.cell(row=1, column=j + 1, value=h)
    ws.cell(row=1, column=n_left + 1, value=c["group_title"])
    ws.merge_cells(start_row=1, start_column=n_left + 1, end_row=1, end_column=total)
    _bold_row(ws, 1, total)
    # 第 2 列：左半是母階資料值（不是欄名！），右半才是群組底下的細項欄名
    for j, v in enumerate(c["parent_row"]):
        ws.cell(row=2, column=j + 1, value=v)
    for j, h in enumerate(c["items"]):
        ws.cell(row=2, column=n_left + 1 + j, value=h).font = Font(bold=True)
    for i, r in enumerate(c["rows"]):
        for j, v in enumerate(r):
            ws.cell(row=3 + i, column=j + 1, value=v)
    for ci in range(1, total + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 14
    if with_result:
        out = wb.create_sheet(f"{c['target_item']}_異常")
        out.append(c["out_cols"])
        for r in c["expected"]:
            out.append(list(r))
    return wb


def gen_two_tier_header(rng: Random, task_id: str):
    for _ in range(40):
        try:
            c = _build_two_tier(rng)
            break
        except AssertionError:
            continue
    else:
        raise RuntimeError("two_tier_header：40 次都建不出有鑑別度的資料")

    cols_txt = "、".join(c["out_cols"])
    instruction = _compose(rng, {
        "full": [
            f"「{c['sheet']}」表的表頭跨兩列：第 1 列是主欄名，右邊「{c['group_title']}」是合併的群組標題，"
            f"群組底下的細項欄名在第 2 列（第 2 列左半放的是母階資料，不是欄名），資料從第 3 列開始。"
            f"請把「{c['target_item']}」不是「{c['sentinel']}」的列挑出來，"
            f"做成新工作表，只留「{cols_txt}」四欄。",

            f"請在「{c['sheet']}」表找出「{c['target_item']}」這一欄"
            f"（它在第 2 列，屬於第 1 列的「{c['group_title']}」群組），"
            f"把值不等於「{c['sentinel']}」的資料列（第 3 列起）挑出來放新工作表，"
            f"欄位為「{cols_txt}」。",
        ],
        "mid": [
            f"「{c['sheet']}」是兩層表頭的表。把「{c['target_item']}」"
            f"不是「{c['sentinel']}」的列做成新工作表，欄位「{cols_txt}」。",

            f"請挑出「{c['sheet']}」表裡「{c['target_item']}」不合格"
            f"（不是「{c['sentinel']}」）的資料列，做成新工作表，欄位「{cols_txt}」"
            f"（注意表頭不只一列）。",
        ],
        "terse": [
            f"「{c['sheet']}」裡「{c['target_item']}」不是「{c['sentinel']}」的挑出來，"
            f"做成新工作表，欄位「{cols_txt}」。",

            f"把「{c['target_item']}」有問題的（非「{c['sentinel']}」）整理到新工作表，"
            f"欄位只要「{cols_txt}」。",

            f"「{c['target_item']}」不合格的做成新表，欄位「{cols_txt}」。",
        ],
    })

    start_wb, goal_wb = _write_two_tier(c, False), _write_two_tier(c, True)
    known = [c["sheet"]]
    n_left = len(c["left"])

    ref = f'''# 推斷：資料表={c["sheet"]}（表頭跨第1~2列、資料第3列起）｜欄位={c["target_item"]}（在第2列，屬「{c["group_title"]}」群組）｜篩選=值≠"{c["sentinel"]}"｜輸出=新工作表
import openpyxl
wb = openpyxl.load_workbook(INPUT_PATH)
ws = wb["{c["sheet"]}"]

DATA_ROW = 3
N_LEFT = {n_left}          # 第 1 列有欄名的欄數；其後才是群組欄
# 有效欄名：左半取第 1 列，群組欄取第 2 列（第 2 列左半是母階資料，不能當欄名）
headers = []
for col in range(1, ws.max_column + 1):
    top = ws.cell(row=1, column=col).value
    sub = ws.cell(row=2, column=col).value
    headers.append(top if col <= N_LEFT else sub)

def col_index(name):
    return headers.index(name)

ti = col_index("{c["target_item"]}")
OUT = {c["out_cols"]!r}
idx = [col_index(h) for h in OUT]

out = wb.create_sheet("{c["target_item"]}_異常")
out.append(OUT)
for row in ws.iter_rows(min_row=DATA_ROW, values_only=True):
    if row[0] is None:
        continue
    v = row[ti]
    if v is None or str(v).strip() == "{c["sentinel"]}":
        continue
    out.append([row[i] for i in idx])
wb.save(OUTPUT_PATH)
'''

    check = {"target_sheets": known,
             "new_sheet": {"known_sheets": known,
                           "headers": c["out_cols"],
                           "rows": [list(r) for r in c["expected"]]}}
    spec = TaskSpec(task_id, "two_tier_header", instruction, check, ref,
                    meta={"variant": "two_tier_header", "item": c["target_item"],
                          "group_title": c["group_title"], "sentinel": c["sentinel"],
                          "n_expected": len(c["expected"])})
    return spec, start_wb, goal_wb


# ======================================================================
# 變體 4：pair_group —— 同鍵多列的主/替配對
# ======================================================================

def _build_pair_group(rng: Random):
    main_tag, alt_tag = rng.choice(_MS_PAIRS)
    tag_col = rng.choice(["M/S", "主替", "料件屬性"])
    n_groups = rng.randint(10, 16)
    n_pairs = rng.randint(3, 5)                 # 有替代料的組數
    pair_at = set(rng.sample(range(n_groups), n_pairs))

    names = _pick_distinct(rng, _PART_NAMES, n_groups)
    headers = ["項次", "料號", "品名", "數量", tag_col, "類別"]
    rows = []
    for g in range(n_groups):
        cat = rng.choice(_PART_CATS)
        qty = rng.randint(1, 8)
        if g in pair_at:
            rows.append([g + 1, _part_no(rng, 1000 + rng.randint(0, 8999)),
                         names[g], qty, main_tag, cat])
            rows.append([g + 1, _part_no(rng, 1000 + rng.randint(0, 8999)),
                         names[g], qty, alt_tag, cat])
        else:
            rows.append([g + 1, _part_no(rng, 1000 + rng.randint(0, 8999)),
                         names[g], qty, None, cat])

    sub = rng.choice(["pick", "sum"])
    if sub == "pick":
        # 有替代料的項次整組取出（主料與替代料兩列都要）
        expected = [r for r in rows if r[0] in {g + 1 for g in pair_at}]
        out_cols = headers
        assert len(expected) == 2 * n_pairs, "pair_group：配對列數不符"
        _assert_discriminates(expected, rows, "pair_group：整組取出等於全表，沒有鑑別度")
    else:
        # 按類別統計數量合計，替代料不重複計入
        agg: dict[str, int] = {}
        naive: dict[str, int] = {}
        for r in rows:
            agg.setdefault(r[5], 0)
            naive.setdefault(r[5], 0)
            naive[r[5]] += r[3]
            if r[4] != alt_tag:
                agg[r[5]] += r[3]
        out_cols = ["類別", "數量合計"]
        expected = [[k, agg[k]] for k in sorted(agg)]
        naive_rows = [[k, naive[k]] for k in sorted(naive)]
        _assert_discriminates(expected, naive_rows,
                              "pair_group：扣不扣替代料結果一樣，沒有鑑別度")

    return {"headers": headers, "rows": rows, "sub": sub, "out_cols": out_cols,
            "expected": expected, "main_tag": main_tag, "alt_tag": alt_tag,
            "tag_col": tag_col, "sheet": rng.choice(_BOM_SHEET_NAMES)}


def _write_pair_group(c: dict, with_result: bool):
    wb = new_wb()
    write_table(wb.create_sheet(c["sheet"]), c["headers"], c["rows"])
    if with_result:
        out = wb.create_sheet("替代料清單" if c["sub"] == "pick" else "類別合計")
        out.append(c["out_cols"])
        for r in c["expected"]:
            out.append(list(r))
    return wb


def gen_pair_group(rng: Random, task_id: str):
    for _ in range(40):
        try:
            c = _build_pair_group(rng)
            break
        except AssertionError:
            continue
    else:
        raise RuntimeError("pair_group：40 次都建不出有鑑別度的資料")

    if c["sub"] == "pick":
        instruction = _compose(rng, {
            "full": [
                f"「{c['sheet']}」表裡同一個「項次」若同時有「{c['main_tag']}」和「{c['alt_tag']}」兩列，"
                f"就表示這個料有替代料（「{c['tag_col']}」欄標示）。"
                f"請把有替代料的項次「整組」（兩列都要）挑出來放到新工作表，欄位與原表相同。",

                f"請找出「{c['sheet']}」表中有替代料的項次——判斷方式是同一個項次在「{c['tag_col']}」欄"
                f"同時出現「{c['main_tag']}」與「{c['alt_tag']}」。把這些項次的所有列（主料和替代料都要）"
                f"複製到新工作表，表頭與原表一致。",
            ],
            "mid": [
                f"「{c['sheet']}」表的「{c['tag_col']}」欄有標「{c['main_tag']}」和「{c['alt_tag']}」。"
                f"請把有替代料的項次整組挑出來放到新工作表，欄位與原表相同。",
            ],
            "terse": [
                f"「{c['sheet']}」裡有替代料的整組挑出來，做成新工作表，欄位照原表。",

                f"把有替代料的項次（含主料那列）複製到新工作表，表頭與原表一致。",
            ],
        })
    else:
        instruction = _compose(rng, {
            "full": [
                f"「{c['sheet']}」表的「{c['tag_col']}」欄標「{c['alt_tag']}」的是替代料，"
                f"和同項次的「{c['main_tag']}」是二擇一、不會同時使用。"
                f"請按「類別」統計「數量」合計（替代料不計入），做成新工作表，"
                f"欄位「類別、數量合計」，依類別名稱由小到大排序。",

                f"請在新工作表做一份「類別 / 數量合計」統計（依類別名稱升冪）。"
                f"注意「{c['sheet']}」表裡「{c['tag_col']}」欄為「{c['alt_tag']}」的列是替代料，"
                f"與同項次的主料重複，不可重複計入數量。",
            ],
            "mid": [
                f"請按「類別」統計「{c['sheet']}」的「數量」合計，做成新工作表，"
                f"欄位「類別、數量合計」，依類別升冪排序"
                f"（注意「{c['tag_col']}」欄，替代料不能重複算）。",
            ],
            "terse": [
                f"「{c['sheet']}」按「類別」統計「數量」合計，做成新工作表，"
                f"欄位「類別、數量合計」，依類別由小到大排序。",

                f"做一張類別數量合計表（新工作表），欄位「類別、數量合計」，依類別升冪。",
            ],
        })

    start_wb, goal_wb = _write_pair_group(c, False), _write_pair_group(c, True)
    known = [c["sheet"]]

    if c["sub"] == "pick":
        body = f'''
tag_i, key_i = col_index("{c["tag_col"]}"), col_index("項次")
data = [list(r) for r in ws.iter_rows(min_row=2, values_only=True) if r[key_i] is not None]

# 先掃一遍找出「同時有主料與替代料」的項次，再整組輸出
tags = {{}}
for row in data:
    tags.setdefault(row[key_i], set()).add(row[tag_i])
paired = {{k for k, v in tags.items() if "{c["main_tag"]}" in v and "{c["alt_tag"]}" in v}}

out = wb.create_sheet("替代料清單")
out.append(headers)
for row in data:
    if row[key_i] in paired:
        out.append(row)
'''
    else:
        body = f'''
tag_i, cat_i, qty_i = col_index("{c["tag_col"]}"), col_index("類別"), col_index("數量")

agg = {{}}
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        continue
    agg.setdefault(row[cat_i], 0)
    if row[tag_i] != "{c["alt_tag"]}":      # 替代料與主料二擇一，不重複計入
        agg[row[cat_i]] += row[qty_i]

out = wb.create_sheet("類別合計")
out.append({c["out_cols"]!r})
for k in sorted(agg):
    out.append([k, agg[k]])
'''

    ref = f'''# 推斷：資料表={c["sheet"]}（表頭第1列）｜欄位={c["tag_col"]}（"{c["main_tag"]}"=主料、"{c["alt_tag"]}"=替代料）｜鍵=項次｜輸出=新工作表
import openpyxl
wb = openpyxl.load_workbook(INPUT_PATH)
ws = wb["{c["sheet"]}"]

headers = [cell.value for cell in ws[1]]
def col_index(name):
    return headers.index(name)
{body}
wb.save(OUTPUT_PATH)
'''

    check = {"target_sheets": known,
             "new_sheet": {"known_sheets": known,
                           "headers": c["out_cols"],
                           "rows": [list(r) for r in c["expected"]]}}
    spec = TaskSpec(task_id, "pair_group", instruction, check, ref,
                    meta={"variant": f"pair_{c['sub']}", "tag_col": c["tag_col"],
                          "main_tag": c["main_tag"], "alt_tag": c["alt_tag"],
                          "n_expected": len(c["expected"])})
    return spec, start_wb, goal_wb


V6_FAMILIES = {
    "dup_header": gen_dup_header,
    "misaligned_merge": gen_misaligned_merge,
    "two_tier_header": gen_two_tier_header,
    "pair_group": gen_pair_group,
}
