"""十二個合成任務家族（schema 泛化版）。

每個家族函式 gen_xxx(rng, task_id) -> (TaskSpec, start_wb, goal_wb)：
- 表格外皮由 schema 輪換（訂單/報銷/庫存/工時/銷售），欄位角色一致
- 指令模板與參考解法都以 token 帶入實際的表名/欄名
- context_rule 家族：關鍵規則只寫在【補充說明】(spec.context)，模型必須讀取才能解對
"""
from __future__ import annotations

from random import Random

from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .base import TaskSpec, build_wb, fill, make_table, schema_tokens

# ----------------------------------------------------------------------
# 參考解法共用片段（__SHEET__ 由 schema 帶入）
# ----------------------------------------------------------------------

_HDR = '''import openpyxl
wb = openpyxl.load_workbook(INPUT_PATH)
ws = wb["__SHEET__"]
headers = [c.value for c in ws[1]]
rows = [list(r) for r in ws.iter_rows(min_row=2, values_only=True)]
'''

_REWRITE = '''if ws.max_row > 1:
    ws.delete_rows(2, ws.max_row - 1)
for r in new_rows:
    ws.append(r)
wb.save(OUTPUT_PATH)
'''


def _ai(meta) -> int:
    return meta["idx"][meta["col"]["amount"]]


def _pick_group(rng: Random, rows, meta):
    """挑一個出現 2 ~ n-2 次的群組值（供刪除類任務）；失敗回 None。"""
    gi = meta["idx"][meta["col"]["group"]]
    counts: dict[str, int] = {}
    for r in rows:
        counts[r[gi]] = counts.get(r[gi], 0) + 1
    cand = [g for g, k in counts.items() if 2 <= k <= len(rows) - 2]
    return rng.choice(sorted(cand)) if cand else None


# ----------------------------------------------------------------------
# 1. filter_rows
# ----------------------------------------------------------------------

def gen_filter_rows(rng: Random, task_id: str):
    variant = rng.choice(["keep_amount", "delete_group"])
    date_col = rng.random() < 0.4

    if variant == "keep_amount":
        headers, rows, meta = make_table(rng, date_col=date_col, unique_amounts=True)
        t = schema_tokens(meta["s"])
        ai = _ai(meta)
        amts = sorted(r[ai] for r in rows)
        thr = amts[len(amts) // 2]
        new_rows = [r for r in rows if r[ai] >= thr]
        instruction = fill(rng.choice([
            "請只保留「__SHEET__」工作表中__AMT__大於或等於 __T__ 的資料列，其餘刪除，欄位與順序維持不變。",
            "把「__SHEET__」裡__AMT__低於 __T__ 的資料列刪掉，只留下__AMT__ ≥ __T__ 的紀錄。",
        ]), t=thr, **t)
        ref = fill(_HDR + '''ai = headers.index("__AMT__")
new_rows = [r for r in rows if r[ai] is not None and r[ai] >= __T__]
''' + _REWRITE, t=thr, **t)
    else:
        for _ in range(30):
            headers, rows, meta = make_table(rng, date_col=date_col, group_pool=3)
            gval = _pick_group(rng, rows, meta)
            if gval:
                break
        t = schema_tokens(meta["s"])
        gi = meta["idx"][meta["col"]["group"]]
        new_rows = [r for r in rows if r[gi] != gval]
        instruction = fill(rng.choice([
            "請刪除「__SHEET__」工作表中__GROUP__為「__GVAL__」的所有資料列，其他列保持原本順序。",
            "「__SHEET__」裡凡是__GROUP__欄等於「__GVAL__」的資料列都不需要了，請整列移除。",
        ]), gval=gval, **t)
        ref = fill(_HDR + '''gi = headers.index("__GROUP__")
new_rows = [r for r in rows if r[gi] != "__GVAL__"]
''' + _REWRITE, gval=gval, **t)

    spec = TaskSpec(task_id, "filter_rows", instruction,
                    {"target_sheets": [t["sheet"]]}, ref,
                    meta={"variant": variant, "schema": meta["s"]["sheet"]})
    return spec, build_wb({t["sheet"]: (headers, rows)}), build_wb({t["sheet"]: (headers, new_rows)})


# ----------------------------------------------------------------------
# 2. sort_rows
# ----------------------------------------------------------------------

def gen_sort_rows(rng: Random, task_id: str):
    desc = rng.random() < 0.6
    headers, rows, meta = make_table(rng, unique_amounts=True, date_col=rng.random() < 0.3)
    t = schema_tokens(meta["s"])
    ai = _ai(meta)
    new_rows = sorted(rows, key=lambda r: r[ai], reverse=desc)
    word = "由大到小" if desc else "由小到大"
    instruction = fill(rng.choice([
        f"請將「__SHEET__」工作表的資料列依「__AMT__」{word}重新排序（標題列保持在第 1 列不動）。",
        f"幫我把「__SHEET__」按__AMT__{word}排好，欄位不變，只調整資料列的順序。",
    ]), **t)
    ref = fill(_HDR + '''ai = headers.index("__AMT__")
new_rows = sorted(rows, key=lambda r: r[ai], reverse=__REV__)
''' + _REWRITE, rev=desc, **t)
    spec = TaskSpec(task_id, "sort_rows", instruction,
                    {"target_sheets": [t["sheet"]]}, ref,
                    meta={"variant": "desc" if desc else "asc", "schema": meta["s"]["sheet"]})
    return spec, build_wb({t["sheet"]: (headers, rows)}), build_wb({t["sheet"]: (headers, new_rows)})


# ----------------------------------------------------------------------
# 3. groupby_summary
# ----------------------------------------------------------------------

def gen_groupby_summary(rng: Random, task_id: str):
    headers, rows, meta = make_table(rng)
    t = schema_tokens(meta["s"])
    ci, ai = meta["idx"][meta["col"]["cat"]], _ai(meta)
    total_h = "總" + t["amt"]
    stats: dict[str, list[int]] = {}
    for r in rows:
        s = stats.setdefault(r[ci], [0, 0])
        s[0] += r[ai]
        s[1] += 1
    sum_rows = [[cat, stats[cat][0], stats[cat][1]] for cat in sorted(stats)]
    instruction = fill(rng.choice([
        "請新增一個名為「彙總」的工作表，統計「__SHEET__」中每個__CAT__的__TOTAL_H__與筆數。"
        "欄位依序為：__CAT__、__TOTAL_H__、筆數，並依__CAT__名稱排序。",
        "幫我做分類統計：建立「彙總」工作表，列出各__CAT__（依名稱排序）的__TOTAL_H__與筆數，"
        "三個欄位標題為 __CAT__、__TOTAL_H__、筆數。",
    ]), total_h=total_h, **t)
    ref = fill(_HDR + '''ci = headers.index("__CAT__")
ai = headers.index("__AMT__")
stats = {}
for r in rows:
    if r[ci] is None:
        continue
    s = stats.setdefault(r[ci], [0, 0])
    s[0] += r[ai]
    s[1] += 1
out = wb.create_sheet("彙總")
out.append(["__CAT__", "__TOTAL_H__", "筆數"])
for cat in sorted(stats):
    out.append([cat, stats[cat][0], stats[cat][1]])
wb.save(OUTPUT_PATH)
''', total_h=total_h, **t)
    spec = TaskSpec(task_id, "groupby_summary", instruction,
                    {"target_sheets": [t["sheet"], "彙總"]}, ref,
                    meta={"schema": meta["s"]["sheet"]})
    start = build_wb({t["sheet"]: (headers, rows)})
    goal = build_wb({t["sheet"]: (headers, rows),
                     "彙總": ([t["cat"], total_h, "筆數"], sum_rows)})
    return spec, start, goal


# ----------------------------------------------------------------------
# 4. compute_column
# ----------------------------------------------------------------------

def gen_compute_column(rng: Random, task_id: str):
    variant = rng.choice(["amount", "discount"])
    if variant == "amount":
        headers, rows, meta = make_table(rng, with_amount=False)
        t = schema_tokens(meta["s"])
        qi, pi = meta["idx"][meta["col"]["qty"]], meta["idx"][meta["col"]["price"]]
        new_rows = [r + [r[qi] * r[pi]] for r in rows]
        new_headers = headers + [t["amt"]]
        instruction = fill(rng.choice([
            "請在「__SHEET__」工作表最右側新增「__AMT__」欄位，內容為 __QTY__ × __PRICE__ 的計算結果（直接填入數值）。",
            "幫「__SHEET__」加上一個「__AMT__」欄（放在最後一欄），值等於每列的__QTY__乘以__PRICE__。",
        ]), **t)
        body = '''qi = headers.index("__QTY__")
pi = headers.index("__PRICE__")
col = len(headers) + 1
ws.cell(row=1, column=col, value="__AMT__")
for i, r in enumerate(rows):
    ws.cell(row=i + 2, column=col, value=r[qi] * r[pi])
wb.save(OUTPUT_PATH)
'''
    else:
        headers, rows, meta = make_table(rng)
        t = schema_tokens(meta["s"])
        pi = meta["idx"][meta["col"]["price"]]
        new_rows = [r + [r[pi] * 9 // 10] for r in rows]
        new_headers = headers + ["折扣價"]
        instruction = fill(rng.choice([
            "請在「__SHEET__」工作表最右側新增「折扣價」欄位：__PRICE__打九折後，小數點以下無條件捨去取整數。",
            "幫「__SHEET__」加一個「折扣價」欄（最後一欄）＝__PRICE__ × 0.9，去掉小數（無條件捨去）。",
        ]), **t)
        body = '''pi = headers.index("__PRICE__")
col = len(headers) + 1
ws.cell(row=1, column=col, value="折扣價")
for i, r in enumerate(rows):
    ws.cell(row=i + 2, column=col, value=r[pi] * 9 // 10)
wb.save(OUTPUT_PATH)
'''
    spec = TaskSpec(task_id, "compute_column", instruction,
                    {"target_sheets": [t["sheet"]]}, fill(_HDR + body, **t),
                    meta={"variant": variant, "schema": meta["s"]["sheet"]})
    return spec, build_wb({t["sheet"]: (headers, rows)}), build_wb({t["sheet"]: (new_headers, new_rows)})


# ----------------------------------------------------------------------
# 5. total_row
# ----------------------------------------------------------------------

def gen_total_row(rng: Random, task_id: str):
    headers, rows, meta = make_table(rng)
    t = schema_tokens(meta["s"])
    qi, ai = meta["idx"][meta["col"]["qty"]], _ai(meta)
    total = [None] * len(headers)
    total[0] = "總計"
    total[qi] = sum(r[qi] for r in rows)
    total[ai] = sum(r[ai] for r in rows)
    instruction = fill(rng.choice([
        "請在「__SHEET__」工作表資料的最底部新增一列總計：第一欄填「總計」，"
        "「__QTY__」與「__AMT__」欄填入該欄的總和（數值），其餘欄位留空。",
        "幫「__SHEET__」加一列總計（放在最後一列）：第一欄寫「總計」，__QTY__與__AMT__兩欄放加總結果，其他欄空白。",
    ]), **t)
    ref = fill(_HDR + '''qi = headers.index("__QTY__")
ai = headers.index("__AMT__")
total = [None] * len(headers)
total[0] = "總計"
total[qi] = sum(r[qi] for r in rows if r[qi] is not None)
total[ai] = sum(r[ai] for r in rows if r[ai] is not None)
ws.append(total)
wb.save(OUTPUT_PATH)
''', **t)
    spec = TaskSpec(task_id, "total_row", instruction, {"target_sheets": [t["sheet"]]}, ref,
                    meta={"schema": meta["s"]["sheet"]})
    return spec, build_wb({t["sheet"]: (headers, rows)}), build_wb({t["sheet"]: (headers, rows + [total])})


# ----------------------------------------------------------------------
# 6. format_style
# ----------------------------------------------------------------------

def gen_format_style(rng: Random, task_id: str):
    variant = rng.choice(["header", "numfmt", "red_low"])

    if variant == "header":
        headers, rows, meta = make_table(rng)
        t = schema_tokens(meta["s"])
        start = build_wb({t["sheet"]: (headers, rows)}, style_header=False)
        goal = build_wb({t["sheet"]: (headers, rows)}, style_header=False)
        for cell in goal[t["sheet"]][1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        last = get_column_letter(len(headers))
        instruction = fill(rng.choice([
            "請將「__SHEET__」工作表的標題列（第 1 列）文字設為粗體，並把該列底色設為 #D9E1F2。",
            "幫「__SHEET__」的第一列（欄位標題）加粗體、底色填 D9E1F2。",
        ]), **t)
        ref = fill('''import openpyxl
from openpyxl.styles import Font, PatternFill
wb = openpyxl.load_workbook(INPUT_PATH)
ws = wb["__SHEET__"]
for cell in ws[1]:
    if cell.value is not None:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
wb.save(OUTPUT_PATH)
''', **t)
        check = {"target_sheets": [t["sheet"]], "format_checks": [
            {"sheet": t["sheet"], "range": f"A1:{last}1",
             "props": {"bold": True, "fill_rgb": "D9E1F2"}}]}

    elif variant == "numfmt":
        headers, rows, meta = make_table(rng)
        t = schema_tokens(meta["s"])
        start = build_wb({t["sheet"]: (headers, rows)})
        goal = build_wb({t["sheet"]: (headers, rows)})
        gws = goal[t["sheet"]]
        fcs = []
        for name in [t["price"], t["amt"]]:
            ci = meta["idx"][name] + 1
            colL = get_column_letter(ci)
            for r in range(2, len(rows) + 2):
                gws.cell(row=r, column=ci).number_format = "#,##0"
            fcs.append({"sheet": t["sheet"], "range": f"{colL}2:{colL}{len(rows) + 1}",
                        "props": {"number_format": "#,##0"}})
        instruction = fill(rng.choice([
            "請將「__SHEET__」工作表中「__PRICE__」與「__AMT__」兩欄的資料儲存格數值格式設為千分位（#,##0）。",
            "幫「__SHEET__」的__PRICE__、__AMT__欄套用千分位數值格式 #,##0（僅資料列，標題不用）。",
        ]), **t)
        ref = fill(_HDR + '''for name in ["__PRICE__", "__AMT__"]:
    ci = headers.index(name) + 1
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=ci).number_format = "#,##0"
wb.save(OUTPUT_PATH)
''', **t)
        check = {"target_sheets": [t["sheet"]], "format_checks": fcs}

    else:  # red_low
        headers, rows, meta = make_table(rng, unique_amounts=True)
        t = schema_tokens(meta["s"])
        ai = _ai(meta)
        amts = sorted(r[ai] for r in rows)
        thr = amts[len(amts) // 2]
        start = build_wb({t["sheet"]: (headers, rows)})
        goal = build_wb({t["sheet"]: (headers, rows)})
        gws = goal[t["sheet"]]
        last = get_column_letter(len(headers))
        fcs = []
        for i, r in enumerate(rows):
            if r[ai] < thr:
                for c in range(1, len(headers) + 1):
                    gws.cell(row=i + 2, column=c).font = Font(color="FF0000")
                fcs.append({"sheet": t["sheet"], "range": f"A{i + 2}:{last}{i + 2}",
                            "props": {"font_rgb": "FF0000"}})
        instruction = fill(rng.choice([
            "請將「__SHEET__」工作表中__AMT__小於 __T__ 的資料列整列字體顏色改為紅色（FF0000）。",
            "「__SHEET__」裡__AMT__低於 __T__ 的那些列，麻煩把整列文字設成紅字（色碼 FF0000）。",
        ]), t=thr, **t)
        ref = fill(_HDR + '''from openpyxl.styles import Font
ai = headers.index("__AMT__")
for i, r in enumerate(rows):
    if r[ai] is not None and r[ai] < __T__:
        for c in range(1, len(headers) + 1):
            ws.cell(row=i + 2, column=c).font = Font(color="FF0000")
wb.save(OUTPUT_PATH)
''', t=thr, **t)
        check = {"target_sheets": [t["sheet"]], "format_checks": fcs}

    spec = TaskSpec(task_id, "format_style", instruction, check, ref,
                    meta={"variant": variant, "schema": meta["s"]["sheet"]})
    return spec, start, goal


# ----------------------------------------------------------------------
# 7. clean_data
# ----------------------------------------------------------------------

def gen_clean_data(rng: Random, task_id: str):
    variant = rng.choice(["dedupe", "strip", "fill_blank"])

    if variant == "dedupe":
        headers, rows, meta = make_table(rng, dup_rows=rng.randint(2, 3))
        t = schema_tokens(meta["s"])
        seen, goal_rows = set(), []
        for r in rows:
            key = tuple(r)
            if key in seen:
                continue
            seen.add(key)
            goal_rows.append(r)
        instruction = fill(rng.choice([
            "「__SHEET__」工作表中有幾列完全重複的資料，請移除重複列，只保留第一次出現的那一列，其餘順序不變。",
            "請幫「__SHEET__」去除重複：內容完全相同的資料列只留第一筆。",
        ]), **t)
        ref = fill(_HDR + '''seen = set()
new_rows = []
for r in rows:
    key = tuple(r)
    if key in seen:
        continue
    seen.add(key)
    new_rows.append(r)
''' + _REWRITE, **t)

    elif variant == "strip":
        headers, rows, meta = make_table(rng, padded_names=rng.randint(3, 5))
        t = schema_tokens(meta["s"])
        ni = meta["idx"][meta["col"]["person"]]
        goal_rows = [list(r) for r in rows]
        for r in goal_rows:
            if isinstance(r[ni], str):
                r[ni] = r[ni].strip()
        instruction = fill(rng.choice([
            "「__SHEET__」工作表的「__PERSON__」欄有些名字前後帶有多餘空白，請將所有__PERSON__名稱去除前後空白。",
            "請清理「__SHEET__」：把「__PERSON__」欄每個儲存格的前後空白字元刪掉。",
        ]), **t)
        ref = fill('''import openpyxl
wb = openpyxl.load_workbook(INPUT_PATH)
ws = wb["__SHEET__"]
headers = [c.value for c in ws[1]]
ni = headers.index("__PERSON__") + 1
for r in range(2, ws.max_row + 1):
    v = ws.cell(row=r, column=ni).value
    if isinstance(v, str):
        ws.cell(row=r, column=ni, value=v.strip())
wb.save(OUTPUT_PATH)
''', **t)

    else:  # fill_blank
        headers, rows, meta = make_table(rng, blank_group=rng.randint(2, 4))
        t = schema_tokens(meta["s"])
        gi = meta["idx"][meta["col"]["group"]]
        goal_rows = [list(r) for r in rows]
        for r in goal_rows:
            if r[gi] is None:
                r[gi] = "未知"
        instruction = fill(rng.choice([
            "「__SHEET__」工作表的「__GROUP__」欄有些儲存格是空的，請將空白的__GROUP__儲存格填入「未知」。",
            "請把「__SHEET__」中__GROUP__欄的空值全部補成「未知」，其他資料不要動。",
        ]), **t)
        ref = fill('''import openpyxl
wb = openpyxl.load_workbook(INPUT_PATH)
ws = wb["__SHEET__"]
headers = [c.value for c in ws[1]]
gi = headers.index("__GROUP__") + 1
for r in range(2, ws.max_row + 1):
    if ws.cell(row=r, column=gi).value in (None, ""):
        ws.cell(row=r, column=gi, value="未知")
wb.save(OUTPUT_PATH)
''', **t)

    spec = TaskSpec(task_id, "clean_data", instruction,
                    {"target_sheets": [t["sheet"]]}, ref,
                    meta={"variant": variant, "schema": meta["s"]["sheet"]})
    return spec, build_wb({t["sheet"]: (headers, rows)}), build_wb({t["sheet"]: (headers, goal_rows)})


# ----------------------------------------------------------------------
# 8. join_lookup
# ----------------------------------------------------------------------

def gen_join_lookup(rng: Random, task_id: str):
    headers, rows, meta = make_table(rng, empty_price_amount=True, consistent_price=True)
    t = schema_tokens(meta["s"])
    qi = meta["idx"][meta["col"]["qty"]]
    pi = meta["idx"][meta["col"]["price"]]
    ai = _ai(meta)
    ii = meta["idx"][meta["col"]["item"]]
    price_rows = [[p, meta["price_map"][p]] for p in sorted(meta["price_map"])]

    goal_rows = [list(r) for r in rows]
    for r in goal_rows:
        unit = meta["price_map"][r[ii]]
        r[pi] = unit
        r[ai] = unit * r[qi]

    instruction = fill(rng.choice([
        "「__SHEET__」工作表的「__PRICE__」與「__AMT__」欄目前是空的。請依「__LOOKUP__」工作表"
        "（__ITEM__ → __PRICE__）填入每列的__PRICE__，並計算 __AMT__ = __QTY__ × __PRICE__（填入數值）。",
        "請用「__LOOKUP__」裡各__ITEM__的__PRICE__，把「__SHEET__」空白的__PRICE__欄補齊，"
        "__AMT__欄也一併算出來（__QTY__乘以__PRICE__）。",
    ]), **t)
    ref = fill('''import openpyxl
wb = openpyxl.load_workbook(INPUT_PATH)
ws = wb["__SHEET__"]
pws = wb["__LOOKUP__"]
price = {}
for r in pws.iter_rows(min_row=2, values_only=True):
    if r[0] is not None:
        price[r[0]] = r[1]
headers = [c.value for c in ws[1]]
item_i = headers.index("__ITEM__") + 1
qty_i = headers.index("__QTY__") + 1
up_i = headers.index("__PRICE__") + 1
amt_i = headers.index("__AMT__") + 1
for r in range(2, ws.max_row + 1):
    p = ws.cell(row=r, column=item_i).value
    if p is None:
        continue
    unit = price.get(p)
    ws.cell(row=r, column=up_i, value=unit)
    ws.cell(row=r, column=amt_i, value=unit * ws.cell(row=r, column=qty_i).value)
wb.save(OUTPUT_PATH)
''', **t)
    spec = TaskSpec(task_id, "join_lookup", instruction,
                    {"target_sheets": [t["sheet"], t["lookup"]]}, ref,
                    meta={"schema": meta["s"]["sheet"]})
    start = build_wb({t["sheet"]: (headers, rows),
                      t["lookup"]: ([t["item"], t["price"]], price_rows)})
    goal = build_wb({t["sheet"]: (headers, goal_rows),
                     t["lookup"]: ([t["item"], t["price"]], price_rows)})
    return spec, start, goal


# ----------------------------------------------------------------------
# 9. split_concat
# ----------------------------------------------------------------------

def gen_split_concat(rng: Random, task_id: str):
    variant = rng.choice(["split_date", "concat_name"])
    if variant == "split_date":
        headers, rows, meta = make_table(rng, date_col=True)
        t = schema_tokens(meta["s"])
        di = meta["idx"][meta["col"]["date"]]
        new_headers = headers + ["年", "月"]
        new_rows = [r + [r[di].year, r[di].month] for r in rows]
        instruction = fill(rng.choice([
            "請依「__DATE_H__」欄，在「__SHEET__」工作表最右側依序新增「年」與「月」兩個欄位，填入對應的年份與月份數值。",
            "幫「__SHEET__」把__DATE_H__拆出來：最右邊加「年」、「月」兩欄（整數），值取自該列的__DATE_H__。",
        ]), **t)
        ref = fill(_HDR + '''di = headers.index("__DATE_H__")
col = len(headers) + 1
ws.cell(row=1, column=col, value="年")
ws.cell(row=1, column=col + 1, value="月")
for i, r in enumerate(rows):
    d = r[di]
    ws.cell(row=i + 2, column=col, value=d.year)
    ws.cell(row=i + 2, column=col + 1, value=d.month)
wb.save(OUTPUT_PATH)
''', **t)
    else:
        headers, rows, meta = make_table(rng, split_name=True)
        t = schema_tokens(meta["s"])
        si, gi = meta["idx"]["姓"], meta["idx"]["名"]
        new_headers = headers + ["姓名"]
        new_rows = [r + [r[si] + r[gi]] for r in rows]
        instruction = fill(rng.choice([
            "請在「__SHEET__」工作表最右側新增「姓名」欄位，內容為「姓」欄加上「名」欄（直接相連，不加空格）。",
            "幫「__SHEET__」加一個「姓名」欄（最後一欄），把姓和名兩欄的文字接在一起。",
        ]), **t)
        ref = fill(_HDR + '''si = headers.index("姓")
gi = headers.index("名")
col = len(headers) + 1
ws.cell(row=1, column=col, value="姓名")
for i, r in enumerate(rows):
    ws.cell(row=i + 2, column=col, value=r[si] + r[gi])
wb.save(OUTPUT_PATH)
''', **t)
    spec = TaskSpec(task_id, "split_concat", instruction,
                    {"target_sheets": [t["sheet"]]}, ref,
                    meta={"variant": variant, "schema": meta["s"]["sheet"]})
    return spec, build_wb({t["sheet"]: (headers, rows)}), build_wb({t["sheet"]: (new_headers, new_rows)})


# ----------------------------------------------------------------------
# 10. top_n
# ----------------------------------------------------------------------

def gen_top_n(rng: Random, task_id: str):
    n_top = rng.randint(3, 5)
    headers, rows, meta = make_table(rng, unique_amounts=True)
    t = schema_tokens(meta["s"])
    ai = _ai(meta)
    top = sorted(rows, key=lambda r: r[ai], reverse=True)[:n_top]
    sheet2 = f"TOP{n_top}"
    instruction = fill(rng.choice([
        "請新增名為「__SHEET2__」的工作表，放入「__SHEET__」中__AMT__最高的前 __N__ 筆資料"
        "（依__AMT__由大到小排列），欄位與「__SHEET__」相同，「__SHEET__」本身不要更動。",
        "幫我建立「__SHEET2__」工作表：從「__SHEET__」挑出__AMT__前 __N__ 高的資料列，由大到小放進去，欄位照舊。",
    ]), sheet2=sheet2, n=n_top, **t)
    ref = fill(_HDR + '''ai = headers.index("__AMT__")
top = sorted(rows, key=lambda r: r[ai], reverse=True)[:__N__]
out = wb.create_sheet("__SHEET2__")
out.append(headers)
for r in top:
    out.append(r)
wb.save(OUTPUT_PATH)
''', n=n_top, sheet2=sheet2, **t)
    spec = TaskSpec(task_id, "top_n", instruction,
                    {"target_sheets": [t["sheet"], sheet2]}, ref,
                    meta={"n": n_top, "schema": meta["s"]["sheet"]})
    start = build_wb({t["sheet"]: (headers, rows)})
    goal = build_wb({t["sheet"]: (headers, rows), sheet2: (headers, top)})
    return spec, start, goal


# ----------------------------------------------------------------------
# 11. composite — 組合式多步任務
# ----------------------------------------------------------------------

def gen_composite(rng: Random, task_id: str):
    variant = rng.choice(["filter_sort_total", "filter_sort", "clean_compute", "filter_group"])

    if variant in ("filter_sort_total", "filter_group"):
        for _ in range(30):
            headers, rows, meta = make_table(rng, unique_amounts=True, group_pool=3)
            gval = _pick_group(rng, rows, meta)
            if gval:
                break
        t = schema_tokens(meta["s"])
        gi = meta["idx"][meta["col"]["group"]]
        kept = [r for r in rows if r[gi] != gval]

    if variant == "filter_sort_total":
        ai, qi = _ai(meta), meta["idx"][meta["col"]["qty"]]
        kept = sorted(kept, key=lambda r: r[ai], reverse=True)
        total = [None] * len(headers)
        total[0] = "總計"
        total[qi] = sum(r[qi] for r in kept)
        total[ai] = sum(r[ai] for r in kept)
        goal_sheets = {t["sheet"]: (headers, kept + [total])}
        check = {"target_sheets": [t["sheet"]]}
        instruction = fill(rng.choice([
            "請對「__SHEET__」工作表依序完成三個操作：(1) 刪除__GROUP__為「__GVAL__」的所有資料列；"
            "(2) 將剩餘資料列依「__AMT__」由大到小排序；(3) 在資料最底部新增總計列——"
            "第一欄填「總計」，「__QTY__」與「__AMT__」欄填入總和，其他欄留空。",
            "幫我整理「__SHEET__」：先把__GROUP__是「__GVAL__」的資料整列刪掉，"
            "接著按__AMT__從高到低排序，最後在底部加一列總計（第一欄寫「總計」，__QTY__與__AMT__放加總，其餘空白）。",
        ]), gval=gval, **t)
        ref = fill(_HDR + '''gi = headers.index("__GROUP__")
ai = headers.index("__AMT__")
qi = headers.index("__QTY__")
kept = [r for r in rows if r[gi] != "__GVAL__"]
kept.sort(key=lambda r: r[ai], reverse=True)
total = [None] * len(headers)
total[0] = "總計"
total[qi] = sum(r[qi] for r in kept if r[qi] is not None)
total[ai] = sum(r[ai] for r in kept if r[ai] is not None)
new_rows = kept + [total]
''' + _REWRITE, gval=gval, **t)

    elif variant == "filter_group":
        cat_i, ai = meta["idx"][meta["col"]["cat"]], _ai(meta)
        total_h = "總" + t["amt"]
        stats: dict[str, list[int]] = {}
        for r in kept:
            s = stats.setdefault(r[cat_i], [0, 0])
            s[0] += r[ai]
            s[1] += 1
        sum_rows = [[cat, stats[cat][0], stats[cat][1]] for cat in sorted(stats)]
        goal_sheets = {t["sheet"]: (headers, kept),
                       "彙總": ([t["cat"], total_h, "筆數"], sum_rows)}
        check = {"target_sheets": [t["sheet"], "彙總"]}
        instruction = fill(rng.choice([
            "請對「__SHEET__」工作表依序完成：(1) 刪除__GROUP__為「__GVAL__」的所有資料列；"
            "(2) 用刪除後剩下的資料建立「彙總」工作表，統計每個__CAT__的__TOTAL_H__與筆數"
            "（欄位依序：__CAT__、__TOTAL_H__、筆數，依__CAT__名稱排序）。",
            "先把「__SHEET__」中__GROUP__為「__GVAL__」的列刪掉，然後根據剩餘資料新增「彙總」工作表："
            "各__CAT__（依名稱排序）一列，欄位為 __CAT__、__TOTAL_H__、筆數。",
        ]), gval=gval, total_h=total_h, **t)
        ref = fill(_HDR + '''gi = headers.index("__GROUP__")
cat_i = headers.index("__CAT__")
ai = headers.index("__AMT__")
new_rows = [r for r in rows if r[gi] != "__GVAL__"]
stats = {}
for r in new_rows:
    s = stats.setdefault(r[cat_i], [0, 0])
    s[0] += r[ai]
    s[1] += 1
out = wb.create_sheet("彙總")
out.append(["__CAT__", "__TOTAL_H__", "筆數"])
for cat in sorted(stats):
    out.append([cat, stats[cat][0], stats[cat][1]])
''' + _REWRITE, gval=gval, total_h=total_h, **t)

    elif variant == "filter_sort":
        headers, rows, meta = make_table(rng, unique_amounts=True)
        t = schema_tokens(meta["s"])
        ai = _ai(meta)
        amts = sorted(r[ai] for r in rows)
        thr = amts[len(amts) // 2]
        kept = sorted([r for r in rows if r[ai] >= thr], key=lambda r: r[ai], reverse=True)
        goal_sheets = {t["sheet"]: (headers, kept)}
        check = {"target_sheets": [t["sheet"]]}
        instruction = fill(rng.choice([
            "請對「__SHEET__」工作表依序完成：(1) 只保留__AMT__大於或等於 __T__ 的資料列；"
            "(2) 將保留的資料列依「__AMT__」由大到小排序。",
            "幫我處理「__SHEET__」：__AMT__低於 __T__ 的列刪掉，剩下的按__AMT__從高到低排好。",
        ]), t=thr, **t)
        ref = fill(_HDR + '''ai = headers.index("__AMT__")
new_rows = [r for r in rows if r[ai] is not None and r[ai] >= __T__]
new_rows.sort(key=lambda r: r[ai], reverse=True)
''' + _REWRITE, t=thr, **t)

    else:  # clean_compute
        headers, rows, meta = make_table(rng, with_amount=False,
                                         padded_names=rng.randint(3, 5),
                                         blank_group=rng.randint(2, 3))
        t = schema_tokens(meta["s"])
        ni = meta["idx"][meta["col"]["person"]]
        gi = meta["idx"][meta["col"]["group"]]
        qi, pi = meta["idx"][meta["col"]["qty"]], meta["idx"][meta["col"]["price"]]
        goal_rows = [list(r) for r in rows]
        for r in goal_rows:
            if isinstance(r[ni], str):
                r[ni] = r[ni].strip()
            if r[gi] is None:
                r[gi] = "未知"
        goal_rows = [r + [r[qi] * r[pi]] for r in goal_rows]
        goal_sheets = {t["sheet"]: (headers + [t["amt"]], goal_rows)}
        check = {"target_sheets": [t["sheet"]]}
        instruction = fill(rng.choice([
            "請對「__SHEET__」工作表依序完成三個清理與計算：(1) 「__PERSON__」欄名稱去除前後空白；"
            "(2) 「__GROUP__」欄空白儲存格填入「未知」；(3) 最右側新增「__AMT__」欄＝__QTY__ × __PRICE__（填數值）。",
            "幫我整理「__SHEET__」：__PERSON__名字的前後空白清掉、__GROUP__空著的補「未知」，"
            "然後在最後加一個「__AMT__」欄，值是__QTY__乘__PRICE__。",
        ]), **t)
        ref = fill(_HDR + '''ni = headers.index("__PERSON__")
gi = headers.index("__GROUP__")
qi = headers.index("__QTY__")
pi = headers.index("__PRICE__")
col = len(headers) + 1
ws.cell(row=1, column=col, value="__AMT__")
for i, r in enumerate(rows):
    if isinstance(r[ni], str):
        ws.cell(row=i + 2, column=ni + 1, value=r[ni].strip())
    if r[gi] is None:
        ws.cell(row=i + 2, column=gi + 1, value="未知")
    ws.cell(row=i + 2, column=col, value=r[qi] * r[pi])
wb.save(OUTPUT_PATH)
''', **t)

    spec = TaskSpec(task_id, "composite", instruction, check, ref,
                    meta={"variant": variant, "schema": meta["s"]["sheet"]})
    return spec, build_wb({t["sheet"]: (headers, rows)}), build_wb(goal_sheets)


# ----------------------------------------------------------------------
# 12. context_rule — 關鍵規則只在【補充說明】裡（客製化 L1 配套）
# ----------------------------------------------------------------------

_DISCOUNT_DISPLAY = {75: "75 折", 80: "8 折", 85: "85 折", 90: "9 折"}


def gen_context_rule(rng: Random, task_id: str):
    variant = rng.choice(["tax", "discount", "fill_default"])

    if variant == "tax":
        headers, rows, meta = make_table(rng)
        t = schema_tokens(meta["s"])
        ai = _ai(meta)
        goal_rows = [r + [(r[ai] * 105 + 50) // 100] for r in rows]
        goal_sheets = {t["sheet"]: (headers + ["含稅價"], goal_rows)}
        instruction = fill(rng.choice([
            "請在「__SHEET__」工作表最右側新增「含稅價」欄位（填入數值）。",
            "幫「__SHEET__」加上「含稅價」欄，放在最後一欄。",
        ]), **t)
        context = fill("本公司規定：含稅價 = __AMT__ × 1.05，計算結果四捨五入取整數。", **t)
        ref = fill(_HDR + '''ai = headers.index("__AMT__")
col = len(headers) + 1
ws.cell(row=1, column=col, value="含稅價")
for i, r in enumerate(rows):
    ws.cell(row=i + 2, column=col, value=(r[ai] * 105 + 50) // 100)
wb.save(OUTPUT_PATH)
''', **t)

    elif variant == "discount":
        pct = rng.choice(sorted(_DISCOUNT_DISPLAY))
        headers, rows, meta = make_table(rng)
        t = schema_tokens(meta["s"])
        pi = meta["idx"][meta["col"]["price"]]
        goal_rows = [r + [r[pi] * pct // 100] for r in rows]
        goal_sheets = {t["sheet"]: (headers + ["折扣價"], goal_rows)}
        instruction = fill(rng.choice([
            "請在「__SHEET__」工作表最右側新增「折扣價」欄位（填入數值）。",
            "幫「__SHEET__」加上「折扣價」欄，放在最後一欄。",
        ]), **t)
        context = fill(f"本公司折扣規定：折扣價一律為__PRICE__打{_DISCOUNT_DISPLAY[pct]}"
                       f"（即 __PRICE__ × {pct} ÷ 100，小數無條件捨去取整數）。", **t)
        ref = fill(_HDR + f'''pi = headers.index("__PRICE__")
col = len(headers) + 1
ws.cell(row=1, column=col, value="折扣價")
for i, r in enumerate(rows):
    ws.cell(row=i + 2, column=col, value=r[pi] * {pct} // 100)
wb.save(OUTPUT_PATH)
''', **t)

    else:  # fill_default
        headers, rows, meta = make_table(rng, blank_group=rng.randint(2, 4))
        t = schema_tokens(meta["s"])
        gi = meta["idx"][meta["col"]["group"]]
        goal_rows = [list(r) for r in rows]
        for r in goal_rows:
            if r[gi] is None:
                r[gi] = t["default"]
        goal_sheets = {t["sheet"]: (headers, goal_rows)}
        instruction = fill(rng.choice([
            "「__SHEET__」工作表的「__GROUP__」欄有些儲存格沒填，請依公司規定補上預設值。",
            "請把「__SHEET__」中__GROUP__欄的空白儲存格補齊（依公司規定的預設值）。",
        ]), **t)
        context = fill("公司規定：「__GROUP__」未填寫時，一律登記為「__DEFAULT__」。", **t)
        ref = fill('''import openpyxl
wb = openpyxl.load_workbook(INPUT_PATH)
ws = wb["__SHEET__"]
headers = [c.value for c in ws[1]]
gi = headers.index("__GROUP__") + 1
for r in range(2, ws.max_row + 1):
    if ws.cell(row=r, column=gi).value in (None, ""):
        ws.cell(row=r, column=gi, value="__DEFAULT__")
wb.save(OUTPUT_PATH)
''', **t)

    spec = TaskSpec(task_id, "context_rule", instruction,
                    {"target_sheets": [t["sheet"]]}, ref, context=context,
                    meta={"variant": variant, "schema": meta["s"]["sheet"]})
    return spec, build_wb({t["sheet"]: (headers, rows)}), build_wb(goal_sheets)


# ----------------------------------------------------------------------
# 13. large_table — 大表格（120~260 列；觀察被截斷，程式碼必須泛化到看不見的列）
# ----------------------------------------------------------------------

def gen_large_table(rng: Random, task_id: str):
    variant = rng.choice(["lt_filter", "lt_groupby", "lt_compute", "lt_total"])
    n = rng.randint(120, 260)

    if variant == "lt_filter":
        headers, rows, meta = make_table(rng, n=n)
        t = schema_tokens(meta["s"])
        ai = _ai(meta)
        amts = sorted(r[ai] for r in rows)
        thr = amts[len(amts) // 2]
        goal_sheets = {t["sheet"]: (headers, [r for r in rows if r[ai] >= thr])}
        instruction = fill(rng.choice([
            "請只保留「__SHEET__」工作表中__AMT__大於或等於 __T__ 的資料列，其餘刪除（表格很長，請務必處理到最後一列）。",
            "把「__SHEET__」裡__AMT__低於 __T__ 的資料列全部刪掉，注意資料有數百列。",
        ]), t=thr, **t)
        ref = fill(_HDR + '''ai = headers.index("__AMT__")
new_rows = [r for r in rows if r[ai] is not None and r[ai] >= __T__]
''' + _REWRITE, t=thr, **t)

    elif variant == "lt_groupby":
        headers, rows, meta = make_table(rng, n=n)
        t = schema_tokens(meta["s"])
        ci, ai = meta["idx"][meta["col"]["cat"]], _ai(meta)
        total_h = "總" + t["amt"]
        stats: dict[str, list[int]] = {}
        for r in rows:
            s = stats.setdefault(r[ci], [0, 0])
            s[0] += r[ai]
            s[1] += 1
        sum_rows = [[c, stats[c][0], stats[c][1]] for c in sorted(stats)]
        goal_sheets = {t["sheet"]: (headers, rows),
                       "彙總": ([t["cat"], total_h, "筆數"], sum_rows)}
        instruction = fill(
            "請新增「彙總」工作表，統計「__SHEET__」中每個__CAT__的__TOTAL_H__與筆數"
            "（欄位：__CAT__、__TOTAL_H__、筆數，依__CAT__排序）。資料有數百列，請統計全部資料。",
            total_h=total_h, **t)
        ref = fill(_HDR + '''ci = headers.index("__CAT__")
ai = headers.index("__AMT__")
stats = {}
for r in rows:
    if r[ci] is None:
        continue
    s = stats.setdefault(r[ci], [0, 0])
    s[0] += r[ai]
    s[1] += 1
out = wb.create_sheet("彙總")
out.append(["__CAT__", "__TOTAL_H__", "筆數"])
for cat in sorted(stats):
    out.append([cat, stats[cat][0], stats[cat][1]])
wb.save(OUTPUT_PATH)
''', total_h=total_h, **t)
        spec = TaskSpec(task_id, "large_table", instruction,
                        {"target_sheets": [t["sheet"], "彙總"]}, ref,
                        meta={"variant": variant, "n": n, "schema": meta["s"]["sheet"]})
        return spec, build_wb({t["sheet"]: (headers, rows)}), build_wb(goal_sheets)

    elif variant == "lt_compute":
        headers, rows, meta = make_table(rng, n=n, with_amount=False)
        t = schema_tokens(meta["s"])
        qi, pi = meta["idx"][meta["col"]["qty"]], meta["idx"][meta["col"]["price"]]
        goal_sheets = {t["sheet"]: (headers + [t["amt"]],
                                    [r + [r[qi] * r[pi]] for r in rows])}
        instruction = fill(
            "請在「__SHEET__」最右側新增「__AMT__」欄位＝__QTY__ × __PRICE__（數值）。表格有數百列，每一列都要算。",
            **t)
        ref = fill(_HDR + '''qi = headers.index("__QTY__")
pi = headers.index("__PRICE__")
col = len(headers) + 1
ws.cell(row=1, column=col, value="__AMT__")
for i, r in enumerate(rows):
    ws.cell(row=i + 2, column=col, value=r[qi] * r[pi])
wb.save(OUTPUT_PATH)
''', **t)

    else:  # lt_total
        headers, rows, meta = make_table(rng, n=n)
        t = schema_tokens(meta["s"])
        qi, ai = meta["idx"][meta["col"]["qty"]], _ai(meta)
        total = [None] * len(headers)
        total[0] = "總計"
        total[qi] = sum(r[qi] for r in rows)
        total[ai] = sum(r[ai] for r in rows)
        goal_sheets = {t["sheet"]: (headers, rows + [total])}
        instruction = fill(
            "請在「__SHEET__」資料最底部新增總計列：第一欄填「總計」，__QTY__與__AMT__欄填總和，其餘留空。"
            "表格有數百列，總和必須涵蓋全部資料。", **t)
        ref = fill(_HDR + '''qi = headers.index("__QTY__")
ai = headers.index("__AMT__")
total = [None] * len(headers)
total[0] = "總計"
total[qi] = sum(r[qi] for r in rows if r[qi] is not None)
total[ai] = sum(r[ai] for r in rows if r[ai] is not None)
ws.append(total)
wb.save(OUTPUT_PATH)
''', **t)

    spec = TaskSpec(task_id, "large_table", instruction,
                    {"target_sheets": [t["sheet"]]}, ref,
                    meta={"variant": variant, "n": n, "schema": meta["s"]["sheet"]})
    return spec, build_wb({t["sheet"]: (headers, rows)}), build_wb(goal_sheets)


# ----------------------------------------------------------------------

FAMILIES = {
    "filter_rows": gen_filter_rows,
    "sort_rows": gen_sort_rows,
    "groupby_summary": gen_groupby_summary,
    "compute_column": gen_compute_column,
    "total_row": gen_total_row,
    "format_style": gen_format_style,
    "clean_data": gen_clean_data,
    "join_lookup": gen_join_lookup,
    "split_concat": gen_split_concat,
    "top_n": gen_top_n,
    "composite": gen_composite,
    "context_rule": gen_context_rule,
    "large_table": gen_large_table,
}

# v2 難度階梯與 v3 版型對地（在檔尾匯入以避免循環依賴問題）
from .families_v2 import V2_FAMILIES  # noqa: E402

FAMILIES.update(V2_FAMILIES)
V1_FAMILY_LIST = [f for f in FAMILIES if f not in V2_FAMILIES]
V2_FAMILY_LIST = list(V2_FAMILIES)

from .families_v3 import V3_FAMILIES  # noqa: E402

FAMILIES.update(V3_FAMILIES)
V3_FAMILY_LIST = list(V3_FAMILIES)
