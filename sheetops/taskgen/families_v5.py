"""v5 任務家族：簡略/隱含指令的理解（terse_intent）。

來源：真實案例。使用者輸入「config1的分析結果(OOB有問題的，ok則不加入)做在新的工作表」，
模型（明確指令下 20/20 正解）卻完全誤解——訓練指令一律點名工作表與欄位，真實使用者不會。

設計原則
--------
1. **資料消歧、不是指令消歧**：指令可以省略，但工作簿內容必須讓答案唯一。
   產生器用 _assert_unique() 主動檢查：狀態欄只有一張表有、SN 是唯一共同欄、
   目標群組值只出現在主檔的群組欄。
2. **槽位揭露**：七個資訊槽（來源表/狀態欄/哨兵值/主檔/鍵/篩選/輸出）隨機揭露一部分，
   其餘留給推斷。分佈：全露 20%、露 4~5 槽 40%、露 2~3 槽 30%、露 1 槽 10%。
3. **括號補充變體**：模仿真實寫法，把揭露的提示塞進「（…）」。
4. **哨兵推斷**：狀態欄是自由文字＋重複出現的哨兵值（ok/正常/良品…），
   模型要學會「重複出現的短值＝正常，其餘＝有問題」。
5. 新工作表由模型自行命名 → 驗證走 check["new_sheet"]（只比內容不比名字）。
"""
from __future__ import annotations

from random import Random

from openpyxl.styles import Font

from .. import zh_data
from .base import TaskSpec, new_wb, write_table

# ----------------------------------------------------------------------
# 資料建構
# ----------------------------------------------------------------------

_WO_PREFIX = ["D56E1953-SB-VN1", "D56E1954-SB-VN1", "D56E2071-SB-TW2", "D56E2088-SB-TW1"]


def _sn(rng: Random, i: int) -> str:
    return f"T1AH2825{i:04X}"


def _write_titled(ws, title, headers, rows):
    """第 1 列是報表標題、第 2 列才是欄位名（仿真實 OOB 表的版型）。"""
    ws.cell(row=1, column=1, value=title).font = Font(bold=True)
    for j, h in enumerate(headers):
        ws.cell(row=2, column=j + 1, value=h).font = Font(bold=True)
    for i, r in enumerate(rows):
        for j, v in enumerate(r):
            ws.cell(row=3 + i, column=j + 1, value=v)


def _make_decoy_data(rng: Random, groups):
    """干擾表的內容只能算一次——起始檔與目標檔必須寫入完全相同的資料。"""
    piv = [["列標籤", "計數"]] + [[g, rng.randint(3, 20)] for g in groups]
    dist = [["項目"] + list(groups)]
    for name in ("外殼", "PCB", "投入量"):
        dist.append([name] + [rng.randint(5, 60) for _ in groups])
    return piv, dist


def _write_decoys(wb, piv, dist):
    """干擾用工作表：結構與任務無關（沒有 SN、沒有狀態欄），模型不該去動它們。"""
    ws = wb.create_sheet("Sheet1")
    for r in piv:
        ws.append(r)
    ws2 = wb.create_sheet("Distribution")
    for r in dist:
        ws2.append(r)


def _assert_unique(master_headers, inspect_headers, status_col, sentinel,
                   group_col, target_group, master_rows, inspect_rows):
    """保證「省略資訊仍可推斷」——不滿足就是壞題，直接拋出讓 selftest 抓到。"""
    # 共同欄位可以不只 SN（真實報表兩邊都有工單號），但只有 SN 值唯一 → 只有它能當鍵
    shared = set(master_headers) & set(inspect_headers)
    assert "SN" in shared, "兩表必須共有 SN"
    for col in shared - {"SN"}:
        vals = [r[master_headers.index(col)] for r in master_rows]
        assert len(set(vals)) < len(vals), f"共同欄位「{col}」值唯一，會與 SN competing 成為鍵"
    assert status_col not in master_headers, "狀態欄不可同時出現在主檔"
    assert group_col not in inspect_headers, "群組欄不可同時出現在檢驗表"
    si = inspect_headers.index(status_col)
    vals = [r[si] for r in inspect_rows]
    n_pass = sum(1 for v in vals if v == sentinel)
    assert n_pass >= 5, f"哨兵值出現次數太少（{n_pass}），無法從資料推斷"
    defects = [v for v in vals if v != sentinel]
    assert len(set(defects)) == len(defects), "缺陷描述必須各不相同（哨兵才能被辨識）"
    gi = master_headers.index(group_col)
    assert any(r[gi] == target_group for r in master_rows), "目標群組在主檔中不存在"


def _build_case(rng: Random):
    """建立一組品管情境資料，回傳所有後續需要的欄位。"""
    n = rng.randint(16, 26)
    sentinel = rng.choice(zh_data.PASS_SENTINELS)
    group_col, group_vals = rng.choice(zh_data.GROUP_SCHEMES)
    target_group = group_vals[0]
    inspect_name = rng.choice(zh_data.INSPECT_SHEET_NAMES)
    master_name = rng.choice(zh_data.MASTER_SHEET_NAMES)
    status_col = rng.choice(["問題點", "問題描述", "檢驗結果", "異常狀況"])
    wo = rng.choice(_WO_PREFIX)

    sns = [_sn(rng, 0x1000 + i) for i in range(n)]
    master_headers = ["No", "工單號", "SN", group_col]
    master_rows = [[i + 1, wo, sns[i], rng.choice(group_vals)] for i in range(n)]

    # 缺陷描述必須各不相同（哨兵才辨識得出來）
    defect_pool = zh_data.DEFECT_TEXTS[:]
    rng.shuffle(defect_pool)
    inspect_headers = ["工單號", "SN", status_col]
    inspect_rows, di = [], 0
    for i in range(n):
        if rng.random() < 0.45 and di < len(defect_pool):
            status = defect_pool[di]
            di += 1
        else:
            status = sentinel
        inspect_rows.append([wo, sns[i], status])

    # 保底：目標群組至少要有 2 筆有問題的，否則答案太少沒鑑別度
    gi = master_headers.index(group_col)
    tgt_idx = [i for i in range(n) if master_rows[i][gi] == target_group]
    if len(tgt_idx) < 4:
        for i in rng.sample(range(n), 4 - len(tgt_idx)):
            master_rows[i][gi] = target_group
        tgt_idx = [i for i in range(n) if master_rows[i][gi] == target_group]
    have = [i for i in tgt_idx if inspect_rows[i][2] != sentinel]
    for i in tgt_idx[:2]:
        if i not in have and di < len(defect_pool):
            inspect_rows[i][2] = defect_pool[di]
            di += 1
    # 哨兵數量保底
    while sum(1 for r in inspect_rows if r[2] == sentinel) < 5:
        for i in range(n):
            if inspect_rows[i][2] != sentinel and i not in tgt_idx:
                inspect_rows[i][2] = sentinel
                break
        else:
            break

    _assert_unique(master_headers, inspect_headers, status_col, sentinel,
                   group_col, target_group, master_rows, inspect_rows)

    group_of = {master_rows[i][2]: master_rows[i][gi] for i in range(n)}
    expected = [list(r) for r in inspect_rows
                if r[2] != sentinel and group_of.get(r[1]) == target_group]

    return dict(n=n, sentinel=sentinel, group_col=group_col, target_group=target_group,
                inspect_name=inspect_name, master_name=master_name, status_col=status_col,
                master_headers=master_headers, master_rows=master_rows,
                inspect_headers=inspect_headers, inspect_rows=inspect_rows,
                expected=expected, groups=group_vals)


# ----------------------------------------------------------------------
# 槽位揭露：組指令
# ----------------------------------------------------------------------

_SLOTS = ("src", "status", "sentinel", "master", "key", "output")   # filter 永遠揭露


def _pick_reveal(rng: Random) -> set[str]:
    """揭露比例分佈：全露 20%、露 4~5 槽 40%、露 2~3 槽 30%、露 1 槽 10%。"""
    roll = rng.random()
    if roll < 0.20:
        k = len(_SLOTS)
    elif roll < 0.60:
        k = rng.choice([4, 5])
    elif roll < 0.90:
        k = rng.choice([2, 3])
    else:
        k = 1
    return set(rng.sample(_SLOTS, k))


def _compose(rng: Random, c: dict, reveal: set[str]) -> str:
    """依揭露的槽位組出自然的繁中指令；部分提示可能被收進括號（模仿真實寫法）。"""
    tg, sc = c["target_group"], c["status_col"]
    paren_style = rng.random() < 0.45

    main = rng.choice([
        f"{tg} 有問題的做一張新表",
        f"把 {tg} 有問題的挑出來，做成新的工作表",
        f"{tg} 的分析結果做在新的工作表",
        f"請把 {tg} 有異常的整理到新工作表",
    ])

    hints = []
    if "src" in reveal:
        hints.append(rng.choice([f"看「{c['inspect_name']}」表", f"資料在「{c['inspect_name']}」"]))
    if "status" in reveal:
        hints.append(rng.choice([f"依「{sc}」欄判斷", f"看「{sc}」欄"]))
    if "sentinel" in reveal:
        hints.append(rng.choice([f"{c['sentinel']} 則不加入", f"{c['sentinel']} 不算問題",
                                 f"排除 {c['sentinel']}"]))
    if "master" in reveal:
        hints.append(rng.choice([f"對照「{c['master_name']}」表的{c['group_col']}",
                                 f"{c['group_col']}在「{c['master_name']}」表"]))
    if "key" in reveal:
        hints.append(rng.choice(["用 SN 對應", "兩表以 SN 串接"]))
    if "output" in reveal:
        hints.append(rng.choice([f"欄位依序為 {'、'.join(c['inspect_headers'])}",
                                 "欄位與原表相同"]))

    if not hints:
        return main
    if paren_style and len(hints) <= 3:
        return f"{main}（{'，'.join(hints)}）"
    return main + "：" + "；".join(hints) + "。"


# ----------------------------------------------------------------------
# 家族本體
# ----------------------------------------------------------------------

def gen_terse_intent(rng: Random, task_id: str):
    for _ in range(40):
        try:
            c = _build_case(rng)
            break
        except AssertionError:
            continue
    else:
        raise RuntimeError("terse_intent：40 次都建不出符合唯一性的資料")

    reveal = _pick_reveal(rng)
    instruction = _compose(rng, c, reveal)

    inspect_title = rng.choice([c["inspect_name"], f"{c['inspect_name']} 紀錄", "檢驗清單"])
    piv, dist = _make_decoy_data(rng, c["groups"])

    def build(with_result: bool):
        wb = new_wb()
        write_table(wb.create_sheet(c["master_name"]), c["master_headers"], c["master_rows"])
        _write_titled(wb.create_sheet(c["inspect_name"]), inspect_title,
                      c["inspect_headers"], c["inspect_rows"])
        _write_decoys(wb, piv, dist)
        if with_result:
            out = wb.create_sheet(f"{c['target_group']}_問題清單")
            out.append(c["inspect_headers"])
            for r in c["expected"]:
                out.append(r)
        return wb

    start_wb, goal_wb = build(False), build(True)
    known = [c["master_name"], c["inspect_name"], "Sheet1", "Distribution"]

    ref = f'''# 推斷：資料表={c["inspect_name"]}（表頭第2列）｜狀態欄={c["status_col"]}（"{c["sentinel"]}"=正常）｜主檔={c["master_name"]}｜鍵=SN｜篩選={c["group_col"]}="{c["target_group"]}"｜輸出=新工作表
import openpyxl
wb = openpyxl.load_workbook(INPUT_PATH)

master = wb["{c["master_name"]}"]
mh = [cell.value for cell in master[1]]
sn_i, grp_i = mh.index("SN"), mh.index("{c["group_col"]}")
group_of = {{}}
for row in master.iter_rows(min_row=2, values_only=True):
    if row[sn_i] is not None:
        group_of[row[sn_i]] = row[grp_i]

insp = wb["{c["inspect_name"]}"]
HEADER_ROW = 2
ih = [cell.value for cell in insp[HEADER_ROW]]
st_i = ih.index("{c["status_col"]}")
isn_i = ih.index("SN")

out = wb.create_sheet("{c["target_group"]}_問題清單")
out.append(ih)
for row in insp.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
    if row[isn_i] is None:
        continue
    status = row[st_i]
    if status is None or str(status).strip() == "{c["sentinel"]}":
        continue
    if group_of.get(row[isn_i]) != "{c["target_group"]}":
        continue
    out.append(list(row))
wb.save(OUTPUT_PATH)
'''

    check = {"target_sheets": known,
             "new_sheet": {"known_sheets": known,
                           "headers": c["inspect_headers"],
                           "rows": c["expected"]}}

    spec = TaskSpec(task_id, "terse_intent", instruction, check, ref,
                    meta={"variant": f"reveal{len(reveal)}",
                          "revealed": sorted(reveal),
                          "sentinel": c["sentinel"],
                          "group": c["target_group"],
                          "n_expected": len(c["expected"])})
    return spec, start_wb, goal_wb


V5_FAMILIES = {
    "terse_intent": gen_terse_intent,
}
