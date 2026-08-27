import openpyxl

wb = openpyxl.load_workbook(INPUT_PATH)
ws = wb["零件清單"]

header_row = 2
col_map = {}
for col in range(1, ws.max_column + 1):
    cell = ws.cell(row=header_row, column=col)
    merged = ws.merged_cells.ranges
    in_merge = any(cell.min_row <= r <= cell.max_row and cell.min_col <= c <= cell.max_col for r, c in merged)
    if in_merge:
        group_title = ws.cell(row=1, column=col).value
        if group_title:
            col_map[group_title] = col
    else:
        val = cell.value
        if val:
            col_map[val] = col

original_col = None
for name, col_idx in col_map.items():
    if name == "原廠":
        original_col = col_idx
        break

if original_col is None:
    raise ValueError("找不到「原廠」欄位")

filtered_data = []
for row in range(3, ws.max_row + 1):
    status = ws.cell(row=row, column=original_col).value
    if status != "正常":
        row_data = []
        for col in range(1, ws.max_column + 1):
            row_data.append(ws.cell(row=row, column=col).value)
        filtered_data.append(row_data)

new_ws = wb.create_sheet(title="原廠異常清單")
for col_idx, name in enumerate(col_map.values(), start=1):
    new_ws.cell(row=1, column=col_idx, value=name)
for row_idx, row_data in enumerate(filtered_data, start=2):
    for col_idx, val in enumerate(row_data, start=1):
        new_ws.cell(row=row_idx, column=col_idx, value=val)

wb.save(OUTPUT_PATH)
