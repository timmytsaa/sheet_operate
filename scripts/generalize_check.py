"""泛化檢查：蒸餾樣本的程式碼換一組資料還對不對。

為什麼需要
----------
Gym 只跑「一個實例」。一段硬編了列數範圍、或假設某區塊固定在第 6~8 欄的程式碼，
在這一題會拿滿分，換一組資料就壞——但它已經被收進訓練資料，示範了錯誤的寫法。

merge_teacher.py 的方法檢查抓的是「看得出來的壞習慣」（字面量 row[N]），
這支抓的是「看不出來但換組資料就露餡」的那種。

做法
----
1. 擾動 start.xlsx：在每張資料表尾端追加 K 列（複製既有列並改值），
   讓任何寫死列數的程式碼露餡。
2. 用「參考解法」跑擾動後的輸入 → 這就是新的正解（參考解法已被 selftest 保證正確）。
3. 用「teacher 的程式碼」跑同一份擾動輸入 → 比對。
4. 不一致 = 該樣本不泛化，剔除。

參考解法本身也會先跑一次原始輸入當 sanity check；若參考解法在擾動後自己就掛了，
表示這個擾動對該家族不適用，該題整個跳過（不冤枉 teacher）。

用法：
  python scripts/generalize_check.py --glob "data/sft/teacher_v6_*.jsonl" --tasks data/tasks/train_v6
  python scripts/generalize_check.py --glob "..." --tasks ... --out data/sft/clean.jsonl
"""
from __future__ import annotations

import argparse
import glob as _glob
import json
import re
import shutil
import sys
import tempfile
from collections import Counter
from pathlib import Path
from random import Random

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from sheetops.executor import run_code          # noqa: E402
from sheetops.verifier import verify            # noqa: E402


def extract(text: str) -> str:
    m = re.search(r"```(?:python)?\n(.*?)```", text, re.S)
    return m.group(1) if m else text


def perturb(src: Path, dst: Path, rng: Random, n_extra: int = 4) -> bool:
    """每張資料表尾端追加幾列（複製既有列、改動可變值）。

    回傳 False 表示這份工作簿不適合擾動（沒有可複製的資料列）。
    """
    wb = openpyxl.load_workbook(src)
    touched = False
    for ws in wb.worksheets:
        if ws.max_row < 3:
            continue
        # 從最後幾列取樣本（避開表頭與可能的群組標題列）
        donors = [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
                  for r in range(max(3, ws.max_row - 5), ws.max_row + 1)]
        donors = [d for d in donors if any(v is not None for v in d)]
        if not donors:
            continue
        for i in range(n_extra):
            base = list(rng.choice(donors))
            for j, v in enumerate(base):
                if isinstance(v, bool):
                    continue
                if isinstance(v, (int, float)):
                    base[j] = v + rng.randint(1, 9)
                elif isinstance(v, str) and v.strip():
                    # 只改「看起來像編號」的字串，避免破壞狀態值/分類值的語意
                    if re.fullmatch(r"[A-Za-z0-9][\w\-./]*\d[\w\-./]*", v.strip()):
                        base[j] = re.sub(r"\d+$", lambda m: str(int(m.group(0)) + 100 + i),
                                         v.strip())
            ws.append(base)
        touched = True
    if touched:
        wb.save(dst)
    return touched


def rebuild_check(check: dict, perturbed_goal: Path) -> dict | None:
    """擾動後正解會變，check 裡寫死的預期列必須依「參考解法在擾動輸入上的輸出」重建。

    verify() 對 new_sheet 是比對 check["new_sheet"]["rows"]（不是比對 goal 工作簿），
    所以只換 goal 檔沒有用——這是這支腳本第一版的 bug，參考解法被誤判 75% 不泛化。
    """
    ns = check.get("new_sheet")
    if not ns:
        return check                      # 沒有 new_sheet 檢查就沿用原本的
    wb = openpyxl.load_workbook(perturbed_goal)
    known = set(ns.get("known_sheets", []))
    added = [s for s in wb.sheetnames if s not in known]
    if len(added) != 1:
        return None                       # 參考解法沒產生剛好一張新表 → 無法判定
    ws = wb[added[0]]
    n_col = len(ns.get("headers") or []) or ws.max_column
    rows = [[ws.cell(r, c).value for c in range(1, n_col + 1)]
            for r in range(2, ws.max_row + 1)]
    new_check = dict(check)
    new_check["new_sheet"] = {**ns, "rows": rows}
    return new_check


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--glob", required=True, help='蒸餾檔樣式，例如 "data/sft/teacher_v6_*.jsonl"')
    ap.add_argument("--tasks", required=True, help="對應的任務目錄")
    ap.add_argument("--out", default=None, help="寫出通過泛化檢查的樣本（省略則只報告）")
    ap.add_argument("--extra-rows", type=int, default=4)
    ap.add_argument("--seed", type=int, default=20260825)
    args = ap.parse_args()

    task_root = Path(args.tasks)
    recs: list[dict] = []
    for fp in sorted(_glob.glob(args.glob)):
        with open(fp, encoding="utf-8") as f:
            recs += [json.loads(l) for l in f if l.strip()]
    if not recs:
        print(f"找不到樣本：{args.glob}")
        return 1
    print(f"樣本 {len(recs)} 筆，任務目錄 {task_root}")

    rng = Random(args.seed)
    by_task: dict[str, dict] = {}
    stat = Counter()
    kept: list[dict] = []
    failures: list[tuple[str, str, str]] = []

    with tempfile.TemporaryDirectory(prefix="genchk_") as tmp:
        tmp = Path(tmp)
        for rec in recs:
            tid = rec["id"]
            # ---- 每題只算一次擾動輸入與新正解 ----
            if tid not in by_task:
                d = task_root / tid
                spec_path = d / "task.json"
                if not spec_path.exists():
                    by_task[tid] = {"skip": "找不到任務目錄"}
                else:
                    spec = json.loads(spec_path.read_text(encoding="utf-8"))
                    pin = tmp / f"{tid}_start.xlsx"
                    if not perturb(d / "start.xlsx", pin, rng, args.extra_rows):
                        by_task[tid] = {"skip": "無法擾動"}
                    else:
                        pgoal = tmp / f"{tid}_goal.xlsx"
                        res = run_code(spec["ref_solution"], pin, pgoal, timeout=40)
                        if not res.ok or not pgoal.exists():
                            # 參考解法自己在擾動後就掛了 → 這個擾動對該家族不適用
                            by_task[tid] = {"skip": "參考解法不耐擾動"}
                        else:
                            chk = rebuild_check(spec["check"], pgoal)
                            if chk is None:
                                by_task[tid] = {"skip": "無法重建預期答案"}
                            else:
                                by_task[tid] = {"in": pin, "goal": pgoal, "check": chk}
            info = by_task[tid]
            if "skip" in info:
                stat[f"跳過：{info['skip']}"] += 1
                kept.append(rec)          # 無法判定就保留，不冤枉
                continue

            out = tmp / f"{tid}_{stat['n']}_out.xlsx"
            stat["n"] += 1
            res = run_code(extract(rec["messages"][-1]["content"]), info["in"], out, timeout=40)
            if not res.ok or not out.exists():
                stat["換資料後執行失敗"] += 1
                failures.append((tid, rec.get("source", "?"),
                                 (res.stderr or "").strip().splitlines()[-1][:70] if res.stderr else "無輸出"))
                continue
            rep = verify(out, info["goal"], info["check"])
            if rep["full_match"]:
                stat["通過"] += 1
                kept.append(rec)
            else:
                stat["換資料後答案不符"] += 1
                failures.append((tid, rec.get("source", "?"),
                                 (rep["mismatches"][:1] or [""])[0][:70]))

    total = stat["通過"] + stat["換資料後執行失敗"] + stat["換資料後答案不符"]
    print(f"\n{'結果':<20}{'筆數':<8}{'佔比'}")
    print("-" * 40)
    for k in ("通過", "換資料後執行失敗", "換資料後答案不符"):
        print(f"{k:<20}{stat[k]:<8}{stat[k] / max(total, 1):.1%}")
    for k in sorted(x for x in stat if x.startswith("跳過")):
        print(f"{k:<20}{stat[k]:<8}（保留，不判定）")

    if failures:
        print(f"\n不泛化的樣本 {len(failures)} 筆（前 8）：")
        src_count = Counter(s for _, s, _ in failures)
        for tid, src, why in failures[:8]:
            print(f"  {tid:<24}{src.split(':')[-1][:16]:<18}{why}")
        print("\n  各 teacher：" + "、".join(f"{s.split(':')[-1]} {n}" for s, n in src_count.most_common()))

    if args.out:
        outp = Path(args.out)
        outp.parent.mkdir(parents=True, exist_ok=True)
        with outp.open("w", encoding="utf-8") as f:
            for r in kept:
                f.write(json.dumps(r, ensure_ascii=False) + "\n")
        print(f"\n寫出通過泛化檢查的 {len(kept)} 筆 → {outp}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
